"""The officer rubric, live scoring, decisions, the second evaluator and the commission
export (spec §8, §9, §10, §10.3 — task 2B).

Two fixtures build every scenario: ``sub4_model`` seeds the real Rev4 criteria (not an empty
test stand-in — the refusal tests need real rubric codes to reject unknown ones against), and
``cycle`` opens a throwaway cycle bound to it so these tests never touch the real seed data.
The one exception is the end-to-end section at the bottom, which loads the real seed on
purpose: brief §2B names Wesa (90.3/A) and Shield (94.7/A) as the acceptance check that
matters most, and that check only means something against the real 13-vendor fixture.
"""

from __future__ import annotations

import uuid
from datetime import UTC, datetime
from typing import Any

import openpyxl
import pytest
from fastapi.testclient import TestClient
from sqlalchemy import select
from sqlalchemy.orm import Session
from vendoriq_api.db import UnitOfWork
from vendoriq_api.models import (
    Application,
    Evaluation,
    QualificationCycle,
    Vendor,
)
from vendoriq_api.models import ScoringModel as ScoringModelRow
from vendoriq_api.models.enums import (
    ApplicationStatus,
    CycleKind,
    CycleStatus,
    ScoringModelStatus,
    UserRole,
    VendorType,
)
from vendoriq_api.seed import real as real_seed
from vendoriq_api.services import applications as applications_service
from vendoriq_scoring import load_model

# ── fixture data ─────────────────────────────────────────────────────────────
#: Every sub-4 criterion at its full value: total 100.0, class A, KO clean.
GOOD_RAW: dict[str, float] = {
    "A.1": 3,
    "A.2": 10,
    "A.3": 3,
    "A.4": 3,
    "B.1": 20_000_000,
    "B.2": 5_000_000,
    "B.3": 3,
    "B.4": 3,
    "C.1": 15,
    "C.2": 8_000_000,
    "C.3": 5,
    "C.4": 3,
    "D.1": 3,
    "D.2": 3,
    "D.3": 3,
    "E.1": 150,
    "E.2": 20,
    "E.3": 3,
    "E.4": 3,
    "F.1": 3,
    "F.2": 3,
    "F.3": 3,
    "G.1": 3,
    "G.2": 10,
}
#: Every non-KO criterion at zero, KO ones just above zero: total 5.7, class F, KO clean —
#: below the 70 pass mark without failing a knock-out (packages/scoring/README.md's rounding).
LOW_RAW: dict[str, float] = {code: (1 if code in ("A.1", "A.4", "F.1") else 0) for code in GOOD_RAW}
#: GOOD_RAW with the licence knock-out at zero: total 95.0, but KO fails.
KO_FAIL_RAW: dict[str, float] = {**GOOD_RAW, "A.1": 0}

_RUBRIC_CODES = [c["code"] for c in load_model("sub-4").criteria if c["kind"] == "rubric"]


def _rubric_subset(raw: dict[str, float]) -> dict[str, int]:
    return {code: int(raw[code]) for code in _RUBRIC_CODES}


# ── fixtures ─────────────────────────────────────────────────────────────────
@pytest.fixture
def sub4_model(session: Session) -> ScoringModelRow:
    """The real Rev4 criteria, loaded into a database row (mirrors ``seed/real.py``)."""
    document = load_model("sub-4")
    row = ScoringModelRow(
        version="sub-4",
        vendor_type=VendorType.SUB,
        name_az=document.name_az,
        name_en=document.name_en,
        status=ScoringModelStatus.ACTIVE,
        groups=list(document.groups),
        criteria=list(document.criteria),
        classes=list(document.classes),
        pass_mark=document.pass_mark,
        validity_months=document.validity_months,
    )
    session.add(row)
    session.commit()
    return row


@pytest.fixture
def cycle(session: Session, sub4_model: ScoringModelRow) -> QualificationCycle:
    row = QualificationCycle(
        name=f"Evaluation test cycle {uuid.uuid4().hex[:6]}",
        kind=CycleKind.TENDER,
        scoring_model_version=sub4_model.version,
        status=CycleStatus.OPEN,
        opens_at=datetime.now(UTC),
    )
    session.add(row)
    session.commit()
    return row


def _application(
    uow: UnitOfWork,
    vendor: Vendor,
    cycle: QualificationCycle,
    *,
    raw: dict[str, float] | None = None,
) -> Application:
    """Invite, run the vendor through submission, and land ``under_review`` (spec §9)."""
    application = applications_service.invite(uow, vendor, cycle_id=cycle.id)
    for step in (
        ApplicationStatus.IN_PROGRESS,
        ApplicationStatus.SUBMITTED,
        ApplicationStatus.UNDER_REVIEW,
    ):
        applications_service.transition(uow, application, step, role=UserRole.OFFICER)
    if raw is not None:
        application.raw_snapshot = dict(raw)
        uow.flush()
    return application


# ── getEvaluation ────────────────────────────────────────────────────────────
def test_get_evaluation_lists_every_criterion_with_raw_value_and_points(
    client: TestClient,
    make_vendor: Any,
    make_user: Any,
    login: Any,
    uow: UnitOfWork,
    cycle: QualificationCycle,
) -> None:
    vendor = make_vendor()
    application = _application(uow, vendor, cycle, raw=GOOD_RAW)
    login(make_user(UserRole.OFFICER))

    response = client.get(f"/api/applications/{application.id}/evaluation")
    assert response.status_code == 200, response.text
    body = response.json()
    assert body["model_version"] == "sub-4"
    assert len(body["rows"]) == 24

    rubric_row = next(row for row in body["rows"] if row["code"] == "A.1")
    assert rubric_row["kind"] == "rubric"
    assert rubric_row["ko"] is True
    assert rubric_row["rubric_score"] == 3
    assert rubric_row["raw_value"] is None
    assert rubric_row["points"] == 5.0

    numeric_row = next(row for row in body["rows"] if row["code"] == "B.1")
    assert numeric_row["kind"] == "thresh"
    assert numeric_row["rubric_score"] is None
    assert numeric_row["raw_value"] == 20_000_000.0

    assert body["computed"]["total"] == 100.0
    assert body["computed"]["cls"] == "A"
    assert body["computed"]["ko"] is True
    assert body["computed"]["pass_mark"] == 70.0
    assert body["can_approve"] is True


# ── putEvaluation ────────────────────────────────────────────────────────────
def test_put_evaluation_saves_and_recomputes(
    client: TestClient,
    make_vendor: Any,
    make_user: Any,
    login: Any,
    uow: UnitOfWork,
    session: Session,
    cycle: QualificationCycle,
) -> None:
    vendor = make_vendor()
    # Numeric raw indicators are already Rev4-perfect; only the rubric cells the officer
    # types in here still need saving — this isolates "recomputes on save" from the numeric
    # side, which `_base_raw` reads straight from the frozen snapshot either way.
    application = _application(uow, vendor, cycle, raw=GOOD_RAW)
    login(make_user(UserRole.OFFICER))

    response = client.put(
        f"/api/applications/{application.id}/evaluation",
        json={"rubric_scores": _rubric_subset(GOOD_RAW), "evidence": {"A.1": "A-04 sighted"}},
    )
    assert response.status_code == 200, response.text
    body = response.json()
    assert body["computed"]["total"] == 100.0
    assert body["computed"]["cls"] == "A"

    session.refresh(application)
    assert application.computed is not None
    assert application.computed["total"] == 100.0
    assert application.rubric_scores == _rubric_subset(GOOD_RAW)

    primary = session.scalar(
        select(Evaluation).where(
            Evaluation.application_id == application.id, Evaluation.is_primary.is_(True)
        )
    )
    assert primary is not None
    assert primary.rubric is not None
    assert primary.rubric["scores"] == _rubric_subset(GOOD_RAW)
    assert primary.rubric["evidence"] == {"A.1": "A-04 sighted"}

    # Lowering the rubric cells and saving again recomputes downward — proves this is a live
    # recalculation, not a value cached from the first save.
    lower = client.put(
        f"/api/applications/{application.id}/evaluation",
        json={"rubric_scores": _rubric_subset(LOW_RAW)},
    )
    assert lower.status_code == 200, lower.text
    assert lower.json()["computed"]["total"] < 100.0


def test_put_evaluation_refuses_unknown_criterion_codes(
    client: TestClient,
    make_vendor: Any,
    make_user: Any,
    login: Any,
    uow: UnitOfWork,
    cycle: QualificationCycle,
) -> None:
    vendor = make_vendor()
    application = _application(uow, vendor, cycle, raw=GOOD_RAW)
    login(make_user(UserRole.OFFICER))

    response = client.put(
        f"/api/applications/{application.id}/evaluation",
        json={"rubric_scores": {"Z.9": 3}},
    )
    assert response.status_code == 422, response.text


def test_a_commission_member_may_not_edit_the_rubric(
    client: TestClient,
    make_vendor: Any,
    make_user: Any,
    login: Any,
    uow: UnitOfWork,
    cycle: QualificationCycle,
) -> None:
    vendor = make_vendor()
    application = _application(uow, vendor, cycle, raw=GOOD_RAW)
    login(make_user(UserRole.COMMISSION))

    response = client.put(
        f"/api/applications/{application.id}/evaluation",
        json={"rubric_scores": _rubric_subset(GOOD_RAW)},
    )
    assert response.status_code == 403, response.text


def test_put_evaluation_refuses_a_retired_model_version(
    client: TestClient,
    make_vendor: Any,
    make_user: Any,
    login: Any,
    uow: UnitOfWork,
    session: Session,
    sub4_model: ScoringModelRow,
    cycle: QualificationCycle,
) -> None:
    vendor = make_vendor()
    application = _application(uow, vendor, cycle, raw=GOOD_RAW)
    sub4_model.status = ScoringModelStatus.RETIRED
    session.commit()
    login(make_user(UserRole.OFFICER))

    response = client.put(
        f"/api/applications/{application.id}/evaluation",
        json={"rubric_scores": _rubric_subset(GOOD_RAW)},
    )
    assert response.status_code == 409, response.text


def test_put_evaluation_is_immutable_once_decided(
    client: TestClient,
    make_vendor: Any,
    make_user: Any,
    login: Any,
    logout: Any,
    uow: UnitOfWork,
    session: Session,
    cycle: QualificationCycle,
) -> None:
    vendor = make_vendor()
    application = _application(uow, vendor, cycle, raw=GOOD_RAW)
    login(make_user(UserRole.OFFICER))
    put = client.put(
        f"/api/applications/{application.id}/evaluation",
        json={"rubric_scores": _rubric_subset(GOOD_RAW)},
    )
    assert put.status_code == 200, put.text
    logout()

    login(make_user(UserRole.MANAGER))
    decide = client.post(f"/api/applications/{application.id}/decide", json={"decision": "approve"})
    assert decide.status_code == 200, decide.text
    logout()

    login(make_user(UserRole.OFFICER))
    again = client.put(
        f"/api/applications/{application.id}/evaluation",
        json={"rubric_scores": _rubric_subset(LOW_RAW)},
    )
    assert again.status_code == 409, again.text


# ── computeScore ─────────────────────────────────────────────────────────────
def test_compute_score_matches_the_engine_and_persists_nothing(
    client: TestClient,
    make_vendor: Any,
    make_user: Any,
    login: Any,
    uow: UnitOfWork,
    session: Session,
    cycle: QualificationCycle,
) -> None:
    vendor = make_vendor()
    application = _application(uow, vendor, cycle, raw=LOW_RAW)
    login(make_user(UserRole.OFFICER))

    response = client.post(
        f"/api/applications/{application.id}/compute",
        json={"rubric_scores": _rubric_subset(GOOD_RAW), "raw_overrides": _numeric_overrides()},
    )
    assert response.status_code == 200, response.text
    assert response.json()["total"] == 100.0
    assert response.json()["cls"] == "A"

    session.refresh(application)
    assert application.computed is None
    assert application.rubric_scores is None


def _numeric_overrides() -> dict[str, float]:
    return {code: value for code, value in GOOD_RAW.items() if code not in _RUBRIC_CODES}


# ── decideApplication ────────────────────────────────────────────────────────
def test_approve_is_refused_below_the_pass_mark(
    client: TestClient,
    make_vendor: Any,
    make_user: Any,
    login: Any,
    logout: Any,
    uow: UnitOfWork,
    cycle: QualificationCycle,
) -> None:
    vendor = make_vendor()
    application = _application(uow, vendor, cycle, raw=LOW_RAW)
    login(make_user(UserRole.OFFICER))
    client.put(
        f"/api/applications/{application.id}/evaluation",
        json={"rubric_scores": _rubric_subset(LOW_RAW)},
    )
    logout()

    login(make_user(UserRole.MANAGER))
    response = client.post(
        f"/api/applications/{application.id}/decide", json={"decision": "approve"}
    )
    assert response.status_code == 409, response.text
    assert response.json()["error"]["details"]["total"] < 70


def test_approve_is_refused_on_a_knock_out_failure(
    client: TestClient,
    make_vendor: Any,
    make_user: Any,
    login: Any,
    logout: Any,
    uow: UnitOfWork,
    cycle: QualificationCycle,
) -> None:
    vendor = make_vendor()
    application = _application(uow, vendor, cycle, raw=KO_FAIL_RAW)
    login(make_user(UserRole.OFFICER))
    client.put(
        f"/api/applications/{application.id}/evaluation",
        json={"rubric_scores": _rubric_subset(KO_FAIL_RAW)},
    )
    logout()

    login(make_user(UserRole.MANAGER))
    response = client.post(
        f"/api/applications/{application.id}/decide", json={"decision": "approve"}
    )
    assert response.status_code == 409, response.text
    assert response.json()["error"]["details"]["ko"] is False


def test_approve_prequalifies_when_ko_passes_above_the_pass_mark(
    client: TestClient,
    make_vendor: Any,
    make_user: Any,
    login: Any,
    logout: Any,
    uow: UnitOfWork,
    session: Session,
    cycle: QualificationCycle,
) -> None:
    vendor = make_vendor()
    application = _application(uow, vendor, cycle, raw=GOOD_RAW)
    login(make_user(UserRole.OFFICER))
    client.put(
        f"/api/applications/{application.id}/evaluation",
        json={"rubric_scores": _rubric_subset(GOOD_RAW)},
    )
    logout()

    login(make_user(UserRole.MANAGER))
    response = client.post(
        f"/api/applications/{application.id}/decide",
        json={"decision": "approve", "valid_months": 6},
    )
    assert response.status_code == 200, response.text
    body = response.json()
    assert body["status"] == "prequalified"
    assert body["decision"] == "approve"
    assert body["cls"] == "A"

    session.refresh(application)
    session.refresh(vendor)
    assert application.decision is not None
    assert application.decision.value == "approve"
    assert application.decided_by is not None
    assert vendor.status.value == "prequalified"


def test_officer_may_not_decide(
    client: TestClient,
    make_vendor: Any,
    make_user: Any,
    login: Any,
    uow: UnitOfWork,
    cycle: QualificationCycle,
) -> None:
    vendor = make_vendor()
    application = _application(uow, vendor, cycle, raw=GOOD_RAW)
    login(make_user(UserRole.OFFICER))
    response = client.post(
        f"/api/applications/{application.id}/decide", json={"decision": "approve"}
    )
    assert response.status_code == 403, response.text


def test_reject_and_request_info_require_a_justification(
    client: TestClient,
    make_vendor: Any,
    make_user: Any,
    login: Any,
    uow: UnitOfWork,
    cycle: QualificationCycle,
) -> None:
    vendor = make_vendor()
    application = _application(uow, vendor, cycle, raw=LOW_RAW)
    login(make_user(UserRole.COMMISSION))

    bare = client.post(f"/api/applications/{application.id}/decide", json={"decision": "reject"})
    assert bare.status_code == 422, bare.text

    justified = client.post(
        f"/api/applications/{application.id}/decide",
        json={"decision": "reject", "justification": "F class, below pass mark."},
    )
    assert justified.status_code == 200, justified.text
    assert justified.json()["status"] == "rejected"
    assert justified.json()["decision"] == "reject"


# ── putSecondEvaluation ──────────────────────────────────────────────────────
def test_second_evaluation_requires_a_primary_first(
    client: TestClient,
    make_vendor: Any,
    make_user: Any,
    login: Any,
    uow: UnitOfWork,
    cycle: QualificationCycle,
) -> None:
    vendor = make_vendor()
    application = _application(uow, vendor, cycle, raw=GOOD_RAW)
    login(make_user(UserRole.COMMISSION))

    response = client.put(
        f"/api/applications/{application.id}/second-evaluator",
        json={"rubric_scores": _rubric_subset(GOOD_RAW)},
    )
    assert response.status_code == 409, response.text


def test_second_evaluation_flags_criteria_that_differ_by_more_than_one_point(
    client: TestClient,
    make_vendor: Any,
    make_user: Any,
    login: Any,
    logout: Any,
    uow: UnitOfWork,
    cycle: QualificationCycle,
) -> None:
    vendor = make_vendor()
    application = _application(uow, vendor, cycle, raw=GOOD_RAW)
    login(make_user(UserRole.OFFICER))
    client.put(
        f"/api/applications/{application.id}/evaluation",
        json={"rubric_scores": _rubric_subset(GOOD_RAW)},
    )
    logout()

    second_rubric = _rubric_subset(GOOD_RAW)
    second_rubric["A.1"] = 1  # differs by 2 from the primary's 3 -> should diverge
    second_rubric["A.3"] = 2  # differs by 1 -> should NOT diverge
    login(make_user(UserRole.COMMISSION))
    response = client.put(
        f"/api/applications/{application.id}/second-evaluator",
        json={"rubric_scores": second_rubric},
    )
    assert response.status_code == 200, response.text
    body = response.json()
    codes = {d["code"] for d in body["divergences"]}
    assert "A.1" in codes
    assert "A.3" not in codes
    entry = next(d for d in body["divergences"] if d["code"] == "A.1")
    assert entry["first"] == 3
    assert entry["second"] == 1


def test_second_evaluator_must_differ_from_the_primary(
    client: TestClient,
    make_vendor: Any,
    make_user: Any,
    login: Any,
    uow: UnitOfWork,
    cycle: QualificationCycle,
) -> None:
    vendor = make_vendor()
    application = _application(uow, vendor, cycle, raw=GOOD_RAW)
    officer = make_user(UserRole.OFFICER)
    login(officer)
    client.put(
        f"/api/applications/{application.id}/evaluation",
        json={"rubric_scores": _rubric_subset(GOOD_RAW)},
    )
    response = client.put(
        f"/api/applications/{application.id}/second-evaluator",
        json={"rubric_scores": _rubric_subset(GOOD_RAW)},
    )
    assert response.status_code == 409, response.text


# ── commission summary export ───────────────────────────────────────────────
def test_commission_summary_exports_reproduce_the_real_rev4_cycle(
    client: TestClient,
    make_user: Any,
    login: Any,
    uow: UnitOfWork,
    session: Session,
    settings: Any,
) -> None:
    real_seed.load_real(uow, settings=settings)
    tqs_cycle = session.scalar(
        select(QualificationCycle).where(QualificationCycle.name == "TQS2026006 Rev4")
    )
    assert tqs_cycle is not None
    login(make_user(UserRole.MANAGER))

    xlsx_response = client.get(f"/api/cycles/{tqs_cycle.id}/export-summary.xlsx")
    assert xlsx_response.status_code == 200, xlsx_response.text
    assert xlsx_response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    import io

    workbook = openpyxl.load_workbook(io.BytesIO(xlsx_response.content))
    assert workbook.sheetnames == ["5. Nəticə Xülasəsi"]
    sheet = workbook.active
    names = {sheet.cell(row=6, column=col).value for col in range(3, sheet.max_column + 1)}
    assert "VVESA MMC (Wesa)" in names
    assert "Shield" in names

    totals: dict[str, float] = {}
    decisions: dict[str, str] = {}
    total_row = None
    result_row = None
    for row in range(13, sheet.max_row + 1):
        label = sheet.cell(row=row, column=2).value
        if isinstance(label, str) and label.startswith("ÜMUMİ BAL"):
            total_row = row
        if isinstance(label, str) and label.startswith("NƏTİCƏ"):
            result_row = row
    assert total_row is not None and result_row is not None
    for col in range(3, sheet.max_column + 1):
        name = sheet.cell(row=6, column=col).value
        if name:
            totals[name] = sheet.cell(row=total_row, column=col).value
            decisions[name] = sheet.cell(row=result_row, column=col).value

    assert totals["VVESA MMC (Wesa)"] == 90.3
    assert decisions["VVESA MMC (Wesa)"].startswith("A —")
    assert totals["Shield"] == 94.7
    assert decisions["Shield"].startswith("A —")

    pdf_response = client.get(f"/api/cycles/{tqs_cycle.id}/export-summary.pdf")
    assert pdf_response.status_code == 200, pdf_response.text
    assert pdf_response.headers["content-type"].startswith("application/pdf")
    assert pdf_response.content.startswith(b"%PDF")
    assert len(pdf_response.content) > 1000


def test_commission_summary_export_requires_staff(client: TestClient) -> None:
    response = client.get(f"/api/cycles/{uuid.uuid4()}/export-summary.xlsx")
    assert response.status_code == 401, response.text


# ── end-to-end: the two vendors the brief names ─────────────────────────────
def test_wesa_and_shield_evaluate_correctly_through_the_endpoints(
    client: TestClient,
    make_user: Any,
    login: Any,
    uow: UnitOfWork,
    session: Session,
    settings: Any,
) -> None:
    """Brief §2B: "an end-to-end check that Wesa evaluates to 90.3/A and Shield to 94.7/A
    through your endpoints is worth more than any synthetic fixture"."""
    real_seed.load_real(uow, settings=settings)
    login(make_user(UserRole.MANAGER))

    for voen, expected_total, expected_cls in (
        ("1003915341", 90.3, "A"),  # VVESA MMC (Wesa)
        ("2002138471", 94.7, "A"),  # Shield
    ):
        vendor = session.scalar(select(Vendor).where(Vendor.voen == voen))
        assert vendor is not None, voen
        application = session.scalar(select(Application).where(Application.vendor_id == vendor.id))
        assert application is not None

        evaluation = client.get(f"/api/applications/{application.id}/evaluation")
        assert evaluation.status_code == 200, evaluation.text
        computed = evaluation.json()["computed"]
        assert computed["total"] == expected_total
        assert computed["cls"] == expected_cls
        assert computed["ko"] is True
        assert evaluation.json()["can_approve"] is True

        # computeScore, called with no overrides, reproduces the same figure the sheet has —
        # this is the same engine call putEvaluation would have made.
        computed_live = client.post(f"/api/applications/{application.id}/compute", json={})
        assert computed_live.status_code == 200, computed_live.text
        assert computed_live.json()["total"] == expected_total
        assert computed_live.json()["cls"] == expected_cls


# ── round 2: three gaps the orchestrator granted routers/vendors.py for ────────
def test_export_vendors_follows_the_current_filter_not_the_whole_register(
    client: TestClient,
    make_user: Any,
    login: Any,
    uow: UnitOfWork,
    session: Session,
    settings: Any,
) -> None:
    """routers/vendors.py, gap 1 of the round-2 report."""
    import io

    real_seed.load_real(uow, settings=settings)
    login(make_user(UserRole.MANAGER))

    unfiltered = client.get("/api/vendors/export.xlsx")
    assert unfiltered.status_code == 200, unfiltered.text
    assert unfiltered.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    workbook = openpyxl.load_workbook(io.BytesIO(unfiltered.content))
    sheet = workbook.active
    rows = list(sheet.iter_rows(min_row=2, values_only=True))
    names = {row[0] for row in rows}
    assert "VVESA MMC (Wesa)" in names
    assert "Shield" in names
    wesa_row = next(row for row in rows if row[0] == "VVESA MMC (Wesa)")
    assert wesa_row[1] == "1003915341"  # VÖEN column
    assert wesa_row[3] == 90.3  # score column
    assert wesa_row[4] == "A"  # class column

    # The filter that follows through: class=A narrows the workbook to just the A-class
    # vendors, same as `listVendors?class=A` narrows the register table.
    filtered = client.get("/api/vendors/export.xlsx", params={"class": "A"})
    assert filtered.status_code == 200, filtered.text
    filtered_names = {
        row[0]
        for row in openpyxl.load_workbook(io.BytesIO(filtered.content)).active.iter_rows(
            min_row=2, values_only=True
        )
    }
    assert filtered_names == {"VVESA MMC (Wesa)", "Shield"}


def test_export_vendors_is_refused_to_a_vendor_caller(
    client: TestClient, make_vendor: Any, make_user: Any, login: Any
) -> None:
    vendor = make_vendor()
    login(make_user(UserRole.VENDOR, vendor=vendor))
    response = client.get("/api/vendors/export.xlsx")
    assert response.status_code == 403, response.text


def test_vendor_detail_carries_the_evaluation_history_across_cycles(
    client: TestClient,
    make_user: Any,
    login: Any,
    uow: UnitOfWork,
    session: Session,
    settings: Any,
) -> None:
    """routers/vendors.py, gap 2: ``evaluations`` was hardcoded to ``[]``."""
    real_seed.load_real(uow, settings=settings)
    wesa = session.scalar(select(Vendor).where(Vendor.voen == "1003915341"))
    assert wesa is not None
    login(make_user(UserRole.OFFICER))

    response = client.get(f"/api/vendors/{wesa.id}")
    assert response.status_code == 200, response.text
    evaluations = response.json()["evaluations"]
    assert len(evaluations) == 1
    entry = evaluations[0]
    assert entry["cycle_name"] == "TQS2026006 Rev4"
    assert entry["model_version"] == "sub-4"
    assert entry["total"] == 90.3
    assert entry["cls"] == "A"
    assert entry["decision"] == "approve"
    assert entry["decided_at"] is not None


def test_vendor_detail_history_includes_an_undecided_application_with_nulls(
    client: TestClient,
    make_vendor: Any,
    make_user: Any,
    login: Any,
    uow: UnitOfWork,
    cycle: QualificationCycle,
) -> None:
    """Unknown stays empty rather than invented: an application still ``under_review`` shows
    up in the history with no total/class/decision, not a fabricated one."""
    vendor = make_vendor()
    application = _application(uow, vendor, cycle, raw=GOOD_RAW)
    login(make_user(UserRole.OFFICER))

    response = client.get(f"/api/vendors/{vendor.id}")
    assert response.status_code == 200, response.text
    evaluations = response.json()["evaluations"]
    assert len(evaluations) == 1
    entry = evaluations[0]
    assert entry["application_id"] == str(application.id)
    assert entry["cycle_name"] == cycle.name
    assert entry["total"] is None
    assert entry["cls"] is None
    assert entry["decision"] is None
    assert entry["decided_at"] is None


def test_decide_response_is_the_consolidated_application_detail(
    client: TestClient,
    make_vendor: Any,
    make_user: Any,
    login: Any,
    logout: Any,
    uow: UnitOfWork,
    cycle: QualificationCycle,
) -> None:
    """routers/vendors.py gap 3 (well, ``services/evaluation.py``): ``decideApplication`` now
    returns the contract's own ``ApplicationDetail`` (task 2A's ``schemas/applications.py``),
    not a locally-defined stand-in. ``score_released`` only exists on the real schema, and its
    value only makes sense as a live computation — a stale duplicate could not get it right."""
    vendor = make_vendor()
    application = _application(uow, vendor, cycle, raw=GOOD_RAW)
    login(make_user(UserRole.OFFICER))
    client.put(
        f"/api/applications/{application.id}/evaluation",
        json={"rubric_scores": _rubric_subset(GOOD_RAW)},
    )
    logout()

    login(make_user(UserRole.MANAGER))
    response = client.post(
        f"/api/applications/{application.id}/decide", json={"decision": "approve"}
    )
    assert response.status_code == 200, response.text
    body = response.json()
    assert body["scoring_model_version"] == "sub-4"
    assert body["computed"]["total"] == 100.0
    assert body["computed"]["model_version"] == "sub-4"
    assert body["rubric_scores"] == _rubric_subset(GOOD_RAW)
    # Score breakdown is released to the vendor only after the commission decision (spec §7) —
    # true here because a decision was just recorded.
    assert body["score_released"] is True
