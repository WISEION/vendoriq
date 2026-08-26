"""Admin: taxonomy, accounts, settings, the audit log and the event stream."""

from __future__ import annotations

import io
from datetime import UTC, datetime, timedelta
from typing import Any

import openpyxl
from fastapi.testclient import TestClient
from sqlalchemy.orm import Session
from vendoriq_api.db import UnitOfWork
from vendoriq_api.models.enums import UserRole
from vendoriq_api.services import settings_store


# ── categories ──────────────────────────────────────────────────────────────
def test_the_taxonomy_round_trips(client: TestClient, make_user: Any, login: Any) -> None:
    login(make_user(UserRole.ADMIN))
    created = client.post(
        "/api/admin/categories",
        json={"code": "flooring", "name_az": "Döşəmə", "name_en": "Flooring", "kind": "work"},
    )
    assert created.status_code == 201, created.text
    category_id = created.json()["id"]

    renamed = client.patch(
        f"/api/admin/categories/{category_id}",
        json={
            "code": "flooring",
            "name_az": "Döşəmə işləri",
            "name_en": "Flooring works",
            "kind": "work",
        },
    )
    assert renamed.json()["name_en"] == "Flooring works"
    assert client.delete(f"/api/admin/categories/{category_id}").status_code == 204


def test_a_duplicate_category_code_is_a_conflict(
    client: TestClient, make_user: Any, make_category: Any, login: Any
) -> None:
    existing = make_category(code="mep")
    login(make_user(UserRole.ADMIN))
    response = client.post(
        "/api/admin/categories",
        json={"code": existing.code, "name_az": "X", "name_en": "X", "kind": "work"},
    )
    assert response.status_code == 409


def test_a_category_in_use_is_deactivated_not_deleted(
    client: TestClient,
    make_user: Any,
    make_category: Any,
    make_vendor: Any,
    login: Any,
    uow: UnitOfWork,
) -> None:
    """Deleting it would leave a vendor's history without a label."""
    from vendoriq_api.services import categories as categories_service

    category = make_category(code="in-use")
    vendor = make_vendor()
    categories_service.set_for_vendor(uow, vendor.id, [category.code])
    uow.commit()

    login(make_user(UserRole.ADMIN))
    assert client.delete(f"/api/admin/categories/{category.id}").status_code == 204
    listed = client.get("/api/admin/categories?include_inactive=true").json()
    row = next(item for item in listed if item["id"] == str(category.id))
    assert row["is_active"] is False


def test_categories_carry_their_vendor_counts(
    client: TestClient,
    make_user: Any,
    make_category: Any,
    make_vendor: Any,
    login: Any,
    uow: UnitOfWork,
) -> None:
    from vendoriq_api.models.enums import VendorStatus
    from vendoriq_api.services import categories as categories_service

    category = make_category(code="counted")
    vendor = make_vendor(status=VendorStatus.PREQUALIFIED)
    categories_service.set_for_vendor(uow, vendor.id, [category.code])
    categories_service.confirm_for_vendor(uow, vendor.id, [category.code])
    uow.commit()

    login(make_user(UserRole.OFFICER))
    row = next(
        item
        for item in client.get("/api/admin/categories").json()
        if item["id"] == str(category.id)
    )
    assert row["vendor_count"] == 1
    assert row["prequalified_count"] == 1


def test_the_taxonomy_is_readable_by_a_vendor(
    client: TestClient, make_user: Any, make_vendor: Any, login: Any
) -> None:
    """A vendor picks its categories from this list, so it has to be able to read it."""
    login(make_user(UserRole.VENDOR, vendor=make_vendor()))
    assert client.get("/api/admin/categories").status_code == 200


# ── users ───────────────────────────────────────────────────────────────────
def test_creating_a_staff_account_returns_the_totp_uri_once(
    client: TestClient, make_user: Any, login: Any
) -> None:
    login(make_user(UserRole.ADMIN))
    created = client.post(
        "/api/admin/users",
        json={
            "email": "New.Officer@vendoriq.test",
            "full_name": "New Officer",
            "role": "officer",
            "password": "Officer!2026",
        },
    )
    assert created.status_code == 201, created.text
    body = created.json()
    assert body["email"] == "new.officer@vendoriq.test"  # normalised
    assert body["has_totp"] is True
    assert body["totp_provisioning_uri"].startswith("otpauth://totp/")

    # The secret is never retrievable again.
    listed = client.get("/api/admin/users?q=new.officer").json()["items"][0]
    assert "totp_provisioning_uri" not in listed


def test_a_vendor_account_needs_a_vendor_id_and_a_staff_account_must_not_have_one(
    client: TestClient, make_user: Any, make_vendor: Any, login: Any
) -> None:
    login(make_user(UserRole.ADMIN))
    vendor = make_vendor()
    missing = client.post("/api/admin/users", json={"email": "portal@example.az", "role": "vendor"})
    assert missing.status_code == 422

    stray = client.post(
        "/api/admin/users",
        json={"email": "officer2@vendoriq.test", "role": "officer", "vendor_id": str(vendor.id)},
    )
    assert stray.status_code == 422

    good = client.post(
        "/api/admin/users",
        json={"email": "portal@example.az", "role": "vendor", "vendor_id": str(vendor.id)},
    )
    assert good.status_code == 201
    assert good.json()["has_totp"] is False  # vendors log in with a one-time code


def test_a_vendor_account_may_not_be_given_a_password(
    client: TestClient, make_user: Any, make_vendor: Any, login: Any
) -> None:
    login(make_user(UserRole.ADMIN))
    response = client.post(
        "/api/admin/users",
        json={
            "email": "pwd@example.az",
            "role": "vendor",
            "vendor_id": str(make_vendor().id),
            "password": "Nope!2026",
        },
    )
    assert response.status_code == 422


def test_a_duplicate_account_email_is_a_conflict(
    client: TestClient, make_user: Any, login: Any
) -> None:
    admin = make_user(UserRole.ADMIN)
    login(admin)
    response = client.post(
        "/api/admin/users", json={"email": admin.email, "role": "officer", "password": "X!2026"}
    )
    assert response.status_code == 409


def test_a_role_change_is_recorded(client: TestClient, make_user: Any, login: Any) -> None:
    login(make_user(UserRole.ADMIN))
    target = make_user(UserRole.OFFICER)
    response = client.put(f"/api/admin/users/{target.id}/role", json={"role": "commission"})
    assert response.status_code == 200
    assert response.json()["role"] == "commission"


def test_the_last_active_admin_cannot_be_demoted_or_deactivated(
    client: TestClient, make_user: Any, login: Any, session: Session
) -> None:
    """Contract: the change that would leave nobody able to undo it is refused."""
    from sqlalchemy import select
    from vendoriq_api.models import User

    # Park every pre-existing admin so exactly one is left standing.
    for other in session.scalars(
        select(User).where(User.role == UserRole.ADMIN, User.is_active.is_(True))
    ):
        other.is_active = False
    session.flush()

    admin = make_user(UserRole.ADMIN)
    login(admin)
    demoted = client.put(f"/api/admin/users/{admin.id}/role", json={"role": "officer"})
    assert demoted.status_code == 409
    assert client.delete(f"/api/admin/users/{admin.id}").status_code == 409


def test_an_admin_may_demote_themself_when_another_admin_remains(
    client: TestClient, make_user: Any, login: Any
) -> None:
    """The guard (`_guard_last_admin`) only refuses the change that would leave nobody able
    to undo it — it does not single out acting on one's own account. Confirmed rather than
    assumed: with a second admin standing, self-demotion is a plain 200."""
    make_user(UserRole.ADMIN)  # a second admin, so the last-admin guard never trips
    acting_admin = make_user(UserRole.ADMIN)
    login(acting_admin)
    response = client.put(f"/api/admin/users/{acting_admin.id}/role", json={"role": "officer"})
    assert response.status_code == 200, response.text
    assert response.json()["role"] == "officer"


def test_deactivating_an_account_keeps_the_row(
    client: TestClient, make_user: Any, login: Any
) -> None:
    """Accounts are deactivated, never deleted — the audit log references them."""
    login(make_user(UserRole.ADMIN))
    target = make_user(UserRole.OFFICER)
    assert client.delete(f"/api/admin/users/{target.id}").status_code == 204
    listed = client.get(f"/api/admin/users?q={target.email}").json()["items"]
    assert listed[0]["is_active"] is False


# ── settings ────────────────────────────────────────────────────────────────
def test_settings_come_back_complete_even_when_nothing_was_stored(
    client: TestClient, make_user: Any, login: Any
) -> None:
    login(make_user(UserRole.MANAGER))
    body = client.get("/api/admin/settings").json()
    assert body["matching"]["strong_min"] == 2
    assert body["matching"]["capacity_ratio"] == 0.40
    assert body["qualification"]["validity_months"] == 12
    assert body["qualification"]["tax_clearance_validity_months"] == 3
    assert body["notifications"]["expiry_reminder_days"] == [30, 7]
    assert body["organisation"]["currency"] == "AZN"


def test_a_partial_update_merges_into_the_group(
    client: TestClient, make_user: Any, login: Any
) -> None:
    login(make_user(UserRole.MANAGER))
    updated = client.put("/api/admin/settings", json={"matching": {"capacity_ratio": 0.5}}).json()
    assert updated["matching"]["capacity_ratio"] == 0.5
    assert updated["matching"]["strong_min"] == 2  # untouched, not dropped


def test_an_unknown_settings_key_or_group_is_rejected(
    client: TestClient, make_user: Any, login: Any
) -> None:
    """A typo that silently created `capacity_ration` would leave matching on the default."""
    login(make_user(UserRole.MANAGER))
    bad_key = client.put("/api/admin/settings", json={"matching": {"capacity_ration": 0.5}})
    bad_group = client.put("/api/admin/settings", json={"matchingg": {"strong_min": 1}})
    assert bad_key.status_code == 422
    assert bad_group.status_code == 422


def test_freshness_windows_convert_months_to_days(session: Session) -> None:
    """Spec §6.6: financials 15 months, headcount 12."""
    windows = settings_store.freshness_windows(session)
    assert windows["B"] == int(15 * 30.4375)
    assert windows["E"] == int(12 * 30.4375)


def test_a_settings_change_is_audited(
    client: TestClient, make_user: Any, login: Any, session: Session
) -> None:
    from sqlalchemy import select
    from vendoriq_api.models import AuditEvent

    login(make_user(UserRole.MANAGER))
    client.put("/api/admin/settings", json={"qualification": {"validity_months": 6}})
    row = session.scalars(
        select(AuditEvent)
        .where(AuditEvent.entity_type == "setting")
        .order_by(AuditEvent.created_at.desc())
    ).first()
    assert row is not None
    assert (row.before or {})["qualification"]["validity_months"] == 12
    assert (row.after or {})["qualification"]["validity_months"] == 6


# ── the audit log ───────────────────────────────────────────────────────────
def test_the_audit_log_lists_the_mutations_with_the_actor(
    client: TestClient, make_vendor: Any, make_user: Any, login: Any
) -> None:
    manager = make_user(UserRole.MANAGER)
    login(manager)
    vendor = make_vendor()
    client.patch(f"/api/vendors/{vendor.id}", json={"region": "Bakı", "reason": "correction"})

    body = client.get(f"/api/admin/audit?entity_type=vendor&entity_id={vendor.id}").json()
    assert body["total"] >= 1
    latest = body["items"][0]
    assert latest["entity_type"] == "vendor"
    assert latest["actor_id"] == str(manager.id)
    assert latest["actor_email"] == manager.email


def test_the_audit_log_filters_by_action_and_actor(
    client: TestClient, make_vendor: Any, make_user: Any, login: Any
) -> None:
    manager = make_user(UserRole.MANAGER)
    login(manager)
    make_vendor()
    by_action = client.get("/api/admin/audit?action=create").json()
    by_actor = client.get(f"/api/admin/audit?actor_id={manager.id}").json()
    assert all(row["action"] == "create" for row in by_action["items"])
    assert all(row["actor_id"] == str(manager.id) for row in by_actor["items"])


# ── audit log export ────────────────────────────────────────────────────────
def test_the_audit_log_export_is_readable_for_committee_minutes(
    client: TestClient, make_user: Any, login: Any
) -> None:
    """Spec §13: "exportable for committee minutes" — a person who was not in the room must
    be able to read it, so before/after must not be a raw JSON dump in a cell."""
    manager = make_user(UserRole.MANAGER)
    login(manager)
    client.put("/api/admin/settings", json={"qualification": {"validity_months": 6}})

    response = client.get("/api/admin/audit/export.xlsx")
    assert response.status_code == 200, response.text
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    sheet = openpyxl.load_workbook(io.BytesIO(response.content)).active
    header = [sheet.cell(row=4, column=col).value for col in range(1, 7)]
    assert header == ["Tarix və vaxt", "İcraçı", "Əməliyyat", "Obyekt", "Əvvəl", "Sonra"]

    # Newest first, same ordering `listAuditEvents` uses — the settings change is row one.
    assert sheet.cell(row=5, column=2).value == manager.email
    assert sheet.cell(row=5, column=3).value == "update"
    assert sheet.cell(row=5, column=4).value == "setting"
    after_cell = sheet.cell(row=5, column=6).value
    assert "qualification.validity_months: 6" in after_cell
    assert "{" not in after_cell and "}" not in after_cell  # no raw JSON


def test_the_audit_log_export_locale_switches_the_headings(
    client: TestClient, make_user: Any, login: Any
) -> None:
    login(make_user(UserRole.ADMIN))
    client.put("/api/admin/settings", json={"matching": {"strong_min": 4}})
    response = client.get("/api/admin/audit/export.xlsx", params={"locale": "en"})
    sheet = openpyxl.load_workbook(io.BytesIO(response.content)).active
    header = [sheet.cell(row=4, column=col).value for col in range(1, 7)]
    assert header == ["Timestamp", "Actor", "Action", "Entity", "Before", "After"]


def test_the_audit_log_export_honours_the_date_window(
    client: TestClient, make_user: Any, login: Any
) -> None:
    login(make_user(UserRole.MANAGER))
    client.put("/api/admin/settings", json={"matching": {"strong_min": 3}})

    future = (datetime.now(UTC) + timedelta(days=1)).isoformat()
    response = client.get("/api/admin/audit/export.xlsx", params={"from": future})
    assert response.status_code == 200, response.text
    sheet = openpyxl.load_workbook(io.BytesIO(response.content)).active
    assert sheet.cell(row=5, column=1).value == "Bu dövr üçün qeyd yoxdur."


def test_the_audit_log_export_is_restricted_to_manager_and_admin(
    client: TestClient, make_user: Any, login: Any
) -> None:
    """Contract: `exportAuditLog` admits `manager` and `admin`, same as `listAuditEvents`."""
    login(make_user(UserRole.OFFICER))
    assert client.get("/api/admin/audit/export.xlsx").status_code == 403


def test_the_audit_log_export_requires_authentication(client: TestClient) -> None:
    assert client.get("/api/admin/audit/export.xlsx").status_code == 401


# ── the event stream ────────────────────────────────────────────────────────
def test_the_event_log_is_readable_and_filterable(
    client: TestClient, make_user: Any, login: Any
) -> None:
    login(make_user(UserRole.OFFICER))
    created = client.post(
        "/api/vendors", json={"legal_name": "Event Probe MMC", "type": "sub"}
    ).json()

    body = client.get(f"/api/events?type=vendor.registered&entity_id={created['id']}").json()
    assert body["total"] == 1
    event = body["items"][0]
    assert event["type"] == "vendor.registered"
    assert event["payload"]["legal_name"] == "Event Probe MMC"


def test_the_event_log_supports_since_for_pollers(
    client: TestClient, make_user: Any, login: Any
) -> None:
    """Brief §2: a future product polls this instead of subscribing to webhooks."""
    login(make_user(UserRole.OFFICER))
    client.post("/api/vendors", json={"legal_name": "Poll One MMC", "type": "sub"})
    marker = client.get("/api/events?page_size=1").json()["items"][0]["created_at"]
    client.post("/api/vendors", json={"legal_name": "Poll Two MMC", "type": "sub"})

    later = client.get(f"/api/events?since={marker}").json()
    names = {item["payload"].get("legal_name") for item in later["items"]}
    assert "Poll Two MMC" in names
    assert "Poll One MMC" not in names


def test_patching_an_account_renames_and_re_passwords_it(
    client: TestClient, make_user: Any, login: Any, session: Session
) -> None:
    login(make_user(UserRole.ADMIN))
    target = make_user(UserRole.OFFICER)
    before = target.password_hash

    response = client.patch(
        f"/api/admin/users/{target.id}",
        json={
            "email": "Renamed.Officer@vendoriq.test",
            "full_name": "Renamed Officer",
            "role": "officer",
            "locale": "en",
            "password": "Renamed!2026",
        },
    )
    assert response.status_code == 200, response.text
    body = response.json()
    assert body["email"] == "renamed.officer@vendoriq.test"
    assert body["full_name"] == "Renamed Officer"
    assert body["locale"] == "en"
    session.refresh(target)
    assert target.password_hash != before


def test_patching_an_account_onto_a_taken_email_is_a_conflict(
    client: TestClient, make_user: Any, login: Any
) -> None:
    admin = make_user(UserRole.ADMIN)
    login(admin)
    target = make_user(UserRole.OFFICER)
    response = client.patch(
        f"/api/admin/users/{target.id}", json={"email": admin.email, "role": "officer"}
    )
    assert response.status_code == 409


def test_a_vendor_account_cannot_be_given_a_password_by_patch(
    client: TestClient, make_user: Any, make_vendor: Any, login: Any
) -> None:
    login(make_user(UserRole.ADMIN))
    portal = make_user(UserRole.VENDOR, vendor=make_vendor())
    response = client.patch(
        f"/api/admin/users/{portal.id}",
        json={"email": portal.email, "role": "vendor", "password": "Nope!2026"},
    )
    assert response.status_code == 422


def test_promoting_a_vendor_account_to_staff_is_refused(
    client: TestClient, make_user: Any, make_vendor: Any, login: Any
) -> None:
    """A staff account must not carry ``vendor_id``; the role change is where that bites."""
    login(make_user(UserRole.ADMIN))
    portal = make_user(UserRole.VENDOR, vendor=make_vendor())
    response = client.put(f"/api/admin/users/{portal.id}/role", json={"role": "officer"})
    assert response.status_code == 422


def test_promoting_to_a_staff_role_enrols_totp(
    client: TestClient, make_user: Any, login: Any, session: Session
) -> None:
    login(make_user(UserRole.ADMIN))
    target = make_user(UserRole.OFFICER, with_totp=False)
    assert target.totp_secret is None
    response = client.put(f"/api/admin/users/{target.id}/role", json={"role": "manager"})
    assert response.json()["has_totp"] is True


def test_an_unknown_user_or_category_is_not_found(
    client: TestClient, make_user: Any, login: Any
) -> None:
    import uuid as _uuid

    login(make_user(UserRole.ADMIN))
    missing = _uuid.uuid4()
    assert client.get("/api/admin/users?q=nobody").status_code == 200
    assert (
        client.patch(
            f"/api/admin/users/{missing}", json={"email": "x@y.az", "role": "officer"}
        ).status_code
        == 404
    )
    assert client.delete(f"/api/admin/categories/{missing}").status_code == 404
