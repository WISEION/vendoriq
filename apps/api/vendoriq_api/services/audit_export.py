"""The audit log export — spec §13: "exportable for committee minutes".

The ``audit_event`` table already carries everything a committee needs (actor, action,
entity, before/after image, timestamp); this module only lays it out for a reader who was
not in the room. That is the whole point of the export, so the one rule that matters here is
the one spec §13 implies by naming the audience: no raw JSON in a cell. ``before``/``after``
are flattened to one ``key: value`` line per changed field instead — readable in Excel, and
readable printed on paper for a signature.

Styled in the spirit of :mod:`vendoriq_api.services.exports` (title band, accent header row)
without importing from it: that module lays out the Rev4 commission-summary sheet, a
different shape with its own participant columns, and this export's only real relative is
the audit table itself.
"""

from __future__ import annotations

from collections.abc import Mapping
from datetime import datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.orm import Session

from ..models import AuditEvent as AuditEventRow
from ..models import User as UserRow

__all__ = ["build_audit_export_workbook"]

#: A generous cap, not a real page — this exports the ``from``/``to`` window, never the whole
#: table (the same convention ``services/exports.py`` uses for the vendor register export).
_EXPORT_LIMIT = 20_000

_NAVY = "1F3864"
_BLUE = "2E5DA3"
_ACCENT = "4472C4"
_WHITE = "FFFFFF"
_LIGHT = "F2F2F2"

_TITLE = {
    "az": "AUDİT JURNALI — KOMİSSİYA PROTOKOLU ÜÇÜN",
    "en": "AUDIT LOG — FOR COMMITTEE MINUTES",
}
_NO_WINDOW = {"az": "Bütün dövr", "en": "Entire period"}
_LABELS: dict[str, dict[str, str]] = {
    "sheet": {"az": "Audit jurnalı", "en": "Audit log"},
    "timestamp": {"az": "Tarix və vaxt", "en": "Timestamp"},
    "actor": {"az": "İcraçı", "en": "Actor"},
    "system": {"az": "Sistem", "en": "System"},
    "action": {"az": "Əməliyyat", "en": "Action"},
    "entity": {"az": "Obyekt", "en": "Entity"},
    "before": {"az": "Əvvəl", "en": "Before"},
    "after": {"az": "Sonra", "en": "After"},
    "empty": {"az": "Bu dövr üçün qeyd yoxdur.", "en": "No entries for this period."},
}


def _label(key: str, locale: str) -> str:
    return _LABELS[key].get(locale, _LABELS[key]["en"])


def _flatten(value: Any, prefix: str = "") -> list[str]:
    """One ``key: value`` line per leaf, the alternative to a raw JSON cell.

    A settings patch touches several groups at once, so a nested dict gets a dotted key
    (``matching.capacity_ratio: 0.5``) and the line still reads as one fact rather than a
    blob a reader has to parse themselves.
    """
    if value is None:
        return []
    if isinstance(value, Mapping):
        lines: list[str] = []
        for key, item in value.items():
            lines.extend(_flatten(item, f"{prefix}.{key}" if prefix else str(key)))
        return lines
    if isinstance(value, list):
        rendered = ", ".join(str(item) for item in value)
        return [f"{prefix}: [{rendered}]" if prefix else f"[{rendered}]"]
    return [f"{prefix}: {value}" if prefix else str(value)]


def _render(value: dict[str, Any] | None) -> str:
    return "\n".join(_flatten(value))


def _window_text(from_: datetime | None, to: datetime | None, locale: str) -> str:
    if from_ is None and to is None:
        return _NO_WINDOW.get(locale, _NO_WINDOW["en"])
    start = from_.strftime("%Y-%m-%d %H:%M") if from_ else "…"
    end = to.strftime("%Y-%m-%d %H:%M") if to else "…"
    return f"{start} → {end}"


def build_audit_export_workbook(
    session: Session,
    *,
    from_: datetime | None = None,
    to: datetime | None = None,
    locale: str = "az",
) -> Workbook:
    """The audit log in the ``from``/``to`` window, newest first — the same ordering
    ``listAuditEvents`` uses (its own docstring: "the committee-minutes view", spec §13).

    Nothing here computes anything: every field is a plain read of the immutable audit row,
    already written by :func:`vendoriq_api.services.audit.record` inside the transaction of
    the mutation it describes.
    """
    query = select(AuditEventRow, UserRow.email).outerjoin(
        UserRow, UserRow.id == AuditEventRow.actor_id
    )
    if from_ is not None:
        query = query.where(AuditEventRow.created_at >= from_)
    if to is not None:
        query = query.where(AuditEventRow.created_at <= to)
    rows = session.execute(
        query.order_by(AuditEventRow.created_at.desc(), AuditEventRow.id.desc()).limit(
            _EXPORT_LIMIT
        )
    ).all()

    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = _label("sheet", locale)[:31]
    sheet.sheet_view.showGridLines = False

    columns = ("timestamp", "actor", "action", "entity", "before", "after")
    widths = (20.0, 26.0, 16.0, 34.0, 52.0, 52.0)
    last_col = len(columns)

    def band(row: int, text: str, *, fill: str, size: float = 13, bold: bool = True) -> None:
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_col)
        cell = sheet.cell(row=row, column=1, value=text)
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.font = Font(name="Calibri", size=size, bold=bold, color=_WHITE)
        cell.alignment = Alignment(horizontal="left", vertical="center")

    band(1, _TITLE.get(locale, _TITLE["en"]), fill=_NAVY, size=14)
    sheet.row_dimensions[1].height = 26
    band(2, _window_text(from_, to, locale), fill=_BLUE, size=10, bold=False)
    sheet.row_dimensions[2].height = 18

    header_row = 4
    for col, key in enumerate(columns, start=1):
        cell = sheet.cell(row=header_row, column=col, value=_label(key, locale))
        cell.font = Font(name="Calibri", size=10, bold=True, color=_WHITE)
        cell.fill = PatternFill("solid", fgColor=_ACCENT)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        sheet.column_dimensions[get_column_letter(col)].width = widths[col - 1]
    sheet.row_dimensions[header_row].height = 20
    sheet.freeze_panes = sheet.cell(row=header_row + 1, column=1)

    data_row = header_row + 1
    for index, (event, actor_email) in enumerate(rows):
        values = (
            event.created_at.strftime("%Y-%m-%d %H:%M:%S"),
            actor_email or _label("system", locale),
            event.action,
            f"{event.entity_type} ({event.entity_id})" if event.entity_id else event.entity_type,
            _render(event.before),
            _render(event.after),
        )
        line_counts = [str(value).count("\n") + 1 for value in values]
        fill = _LIGHT if index % 2 else _WHITE
        for col, value in enumerate(values, start=1):
            cell = sheet.cell(row=data_row, column=col, value=value)
            cell.font = Font(name="Calibri", size=9)
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            cell.fill = PatternFill("solid", fgColor=fill)
        sheet.row_dimensions[data_row].height = min(200.0, max(15.0, max(line_counts) * 14.0))
        data_row += 1

    if not rows:
        sheet.merge_cells(start_row=data_row, start_column=1, end_row=data_row, end_column=last_col)
        empty_cell = sheet.cell(row=data_row, column=1, value=_label("empty", locale))
        empty_cell.font = Font(name="Calibri", size=10, italic=True, color="808080")
        empty_cell.alignment = Alignment(horizontal="left", vertical="center")

    return workbook
