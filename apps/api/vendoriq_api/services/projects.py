"""Projects, work packages and matching runs: repository queries and mutations (spec §11).

``run_match`` is the one function that talks to the scoring/matching engine, and it does so
exclusively through ``services/matching.py`` — this module never computes a threshold, a
capacity ratio or a gap (packages/scoring/README.md, CONTRIBUTING). What it owns is CRUD on
``Project`` / ``WorkPackage`` and persisting the engine's result as a ``MatchRun`` row, so
``getLatestMatch`` shows a manager the same numbers the commission saw (spec §11.2).
"""

from __future__ import annotations

import uuid
from collections.abc import Sequence
from dataclasses import dataclass
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import func, select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from ..db import UnitOfWork
from ..errors import ApiError
from ..models import MatchRun as MatchRunRow
from ..models import Project, WorkPackage
from ..models.enums import EventType, ProjectStage, ScoreClass
from . import audit, categories, events, matching

PROJECT_FIELDS = ("code", "name", "client", "stage", "estimated_value", "deadline", "external_ref")
PROJECT_AUDIT_FIELDS = (*PROJECT_FIELDS, "is_demo")
PACKAGE_AUDIT_FIELDS = ("name", "estimated_value", "min_class", "required_certs", "notes")


@dataclass(frozen=True, slots=True)
class ProjectFilters:
    """Screen 22's controls."""

    stages: Sequence[ProjectStage] = ()
    q: str | None = None
    include_demo: bool = True


def get(session: Session, project_id: uuid.UUID) -> Project:
    project = session.get(Project, project_id)
    if project is None:
        raise ApiError(404, "not_found", "No such project.")
    return project


def by_code(session: Session, code: str) -> Project | None:
    return session.scalar(select(Project).where(Project.code == code))


def package_count(session: Session, project_id: uuid.UUID) -> int:
    return (
        session.scalar(
            select(func.count())
            .select_from(WorkPackage)
            .where(WorkPackage.project_id == project_id)
        )
        or 0
    )


def latest_match(session: Session, project_id: uuid.UUID) -> MatchRunRow | None:
    return session.scalars(
        select(MatchRunRow)
        .where(MatchRunRow.project_id == project_id)
        .order_by(MatchRunRow.ran_at.desc())
        .limit(1)
    ).first()


def _filtered(filters: ProjectFilters) -> Any:
    query = select(Project)
    if filters.stages:
        query = query.where(Project.stage.in_(list(filters.stages)))
    if not filters.include_demo:
        query = query.where(Project.is_demo.is_(False))
    if filters.q:
        needle = f"%{filters.q.strip().lower()}%"
        query = query.where(
            func.lower(Project.code).like(needle) | func.lower(Project.name).like(needle)
        )
    return query


def list_page(
    session: Session, filters: ProjectFilters, *, page: int = 1, page_size: int = 25
) -> tuple[list[Project], int]:
    query = _filtered(filters)
    total = session.scalar(select(func.count()).select_from(query.subquery())) or 0
    rows = session.scalars(
        query.order_by(Project.created_at.desc()).limit(page_size).offset((page - 1) * page_size)
    ).all()
    return list(rows), total


def create(uow: UnitOfWork, data: dict[str, Any]) -> Project:
    code = str(data["code"]).strip()
    if by_code(uow.session, code) is not None:
        raise ApiError(409, "conflict", "A project with this code already exists.", {"code": code})
    project = Project(
        code=code,
        name=str(data["name"]).strip(),
        client=data.get("client"),
        stage=data.get("stage") or ProjectStage.PIPELINE,
        estimated_value=data.get("estimated_value"),
        deadline=data.get("deadline"),
        external_ref=data.get("external_ref"),
        is_demo=bool(data.get("is_demo", False)),
    )
    uow.session.add(project)
    try:
        uow.flush()
    except IntegrityError as exc:
        uow.session.rollback()
        raise ApiError(409, "conflict", "A project with this code already exists.") from exc
    audit.record(
        uow,
        entity_type="project",
        entity_id=project.id,
        action="create",
        after=audit.snapshot(project, PROJECT_AUDIT_FIELDS),
    )
    return project


def patch(uow: UnitOfWork, project: Project, data: dict[str, Any]) -> Project:
    before = audit.snapshot(project, PROJECT_AUDIT_FIELDS)
    if "code" in data and data["code"] and data["code"] != project.code:
        clash = by_code(uow.session, str(data["code"]))
        if clash is not None and clash.id != project.id:
            raise ApiError(409, "conflict", "A project with this code already exists.")
    for key in PROJECT_FIELDS:
        if key not in data or data[key] is None:
            continue
        setattr(project, key, data[key])
    if "is_demo" in data and data["is_demo"] is not None:
        project.is_demo = bool(data["is_demo"])
    uow.flush()
    after = audit.snapshot(project, PROJECT_AUDIT_FIELDS)
    audit.record(
        uow,
        entity_type="project",
        entity_id=project.id,
        action="update",
        before={key: before[key] for key in audit.diff(before, after)},
        after=audit.diff(before, after),
    )
    return project


def delete(uow: UnitOfWork, project: Project) -> None:
    """Deletes the project and its packages/match runs (``ondelete=CASCADE``, contract)."""
    audit.record(
        uow,
        entity_type="project",
        entity_id=project.id,
        action="delete",
        before=audit.snapshot(project, PROJECT_AUDIT_FIELDS),
    )
    uow.session.delete(project)
    uow.flush()


# ── work packages ───────────────────────────────────────────────────────────
def list_packages(session: Session, project_id: uuid.UUID) -> list[WorkPackage]:
    return list(
        session.scalars(
            select(WorkPackage)
            .where(WorkPackage.project_id == project_id)
            .order_by(WorkPackage.created_at.asc())
        )
    )


def get_package(session: Session, project_id: uuid.UUID, package_id: uuid.UUID) -> WorkPackage:
    package = session.get(WorkPackage, package_id)
    if package is None or package.project_id != project_id:
        raise ApiError(404, "not_found", "No such package on this project.")
    return package


def create_package(uow: UnitOfWork, project: Project, data: dict[str, Any]) -> WorkPackage:
    category = categories.by_code(uow.session, str(data["category_code"]))
    if category is None:
        raise ApiError(
            422,
            "validation_error",
            f"Unknown category code {data['category_code']!r}.",
            {"category_code": data["category_code"]},
        )
    package = WorkPackage(
        project_id=project.id,
        category_id=category.id,
        name=data.get("name"),
        estimated_value=data.get("estimated_value") or 0,
        min_class=data.get("min_class") or ScoreClass.C,
        required_certs=list(data.get("required_certs") or []),
        notes=data.get("notes"),
    )
    uow.session.add(package)
    uow.flush()
    audit.record(
        uow,
        entity_type="work_package",
        entity_id=package.id,
        action="create",
        after={
            "project_id": str(project.id),
            "category": category.code,
            **audit.snapshot(package, PACKAGE_AUDIT_FIELDS),
        },
    )
    return package


def patch_package(uow: UnitOfWork, package: WorkPackage, data: dict[str, Any]) -> WorkPackage:
    before = audit.snapshot(package, PACKAGE_AUDIT_FIELDS)
    category_code = data.get("category_code")
    if category_code:
        category = categories.by_code(uow.session, str(category_code))
        if category is None:
            raise ApiError(
                422,
                "validation_error",
                f"Unknown category code {category_code!r}.",
                {"category_code": category_code},
            )
        package.category_id = category.id
    for key in ("name", "estimated_value", "min_class", "required_certs", "notes"):
        if key not in data or data[key] is None:
            continue
        setattr(package, key, data[key])
    uow.flush()
    after = audit.snapshot(package, PACKAGE_AUDIT_FIELDS)
    audit.record(
        uow,
        entity_type="work_package",
        entity_id=package.id,
        action="update",
        before={key: before[key] for key in audit.diff(before, after)},
        after=audit.diff(before, after),
    )
    return package


def delete_package(uow: UnitOfWork, package: WorkPackage) -> None:
    audit.record(
        uow,
        entity_type="work_package",
        entity_id=package.id,
        action="delete",
        before=audit.snapshot(package, PACKAGE_AUDIT_FIELDS),
    )
    uow.session.delete(package)
    uow.flush()


# ── matching ─────────────────────────────────────────────────────────────────
def run_match(
    uow: UnitOfWork, project: Project, params_override: dict[str, float | int | None] | None
) -> MatchRunRow:
    """Run the engine and persist the result — never recomputed by ``getLatestMatch`` (§11.2)."""
    params = matching.resolve_params(uow.session, params_override)
    result = matching.run(uow.session, project, params)

    row = MatchRunRow(
        project_id=project.id,
        params={
            "strong_min": params.strong_min,
            "capacity_ratio": params.capacity_ratio,
            "supplier_turnover_divisor": params.supplier_turnover_divisor,
        },
        result=matching.serialize(result),
        ran_by=uow.actor_id if isinstance(uow.actor_id, uuid.UUID) else None,
    )
    uow.session.add(row)
    uow.flush()
    audit.record(
        uow,
        entity_type="match_run",
        entity_id=row.id,
        action="run",
        after={
            "project_id": str(project.id),
            "state": result.state,
            "coverage_pct": result.coverage_pct,
        },
    )
    events.emit(
        uow,
        EventType.PROJECT_MATCHED,
        entity_type="project",
        entity_id=project.id,
        payload={
            "match_run_id": str(row.id),
            "state": result.state,
            "coverage_pct": result.coverage_pct,
        },
    )
    return row


#: Bilingual column headers (spec §13: every user-facing string is AZ + EN). Kept local to
#: the export because there is no server-side i18n dictionary — only the web app carries
#: one — and the wording matches the vocabulary already approved in the prototype
#: (`docs/design/app.js` `th_value`, `th_class`, `m_cands`, `m_gap`, `m_fit`).
_HEADERS = {
    "az": {
        "package": "İş paketi",
        "value": "Dəyər (AZN)",
        "min_class": "Min. sinif",
        "certs": "Tələb olunan sertifikatlar",
        "state": "Go / No-Go",
        "gap": "Boşluq",
        "vendor": "Vendor",
        "score": "Bal",
        "class": "Sinif",
        "capacity": "Kapasitet uyğunluğu",
        "eligible": "Uyğun",
    },
    "en": {
        "package": "Work package",
        "value": "Value (AZN)",
        "min_class": "Min. class",
        "certs": "Required certificates",
        "state": "Go / No-Go",
        "gap": "Gap",
        "vendor": "Vendor",
        "score": "Score",
        "class": "Class",
        "capacity": "Capacity fit",
        "eligible": "Eligible",
    },
}


def export_workbook(session: Session, project: Project, *, locale: str = "az") -> bytes:
    """Packages and the latest matching result as an Excel workbook (``exportProject``).

    Renders whatever is on file, honestly: a package's ``state``/``gap``/candidate columns
    stay blank when the project has never been matched, rather than showing a stale or
    invented number (brief §4.3 — unknown facts stay empty).
    """
    labels = _HEADERS["en"] if locale == "en" else _HEADERS["az"]
    run = latest_match(session, project.id)
    by_package_id: dict[str, dict[str, Any]] = {}
    if run is not None:
        for package_result in run.result.get("packages", []):
            by_package_id[str(package_result["package_id"])] = package_result

    workbook = Workbook()
    packages_sheet = workbook.active
    assert packages_sheet is not None
    packages_sheet.title = "Packages"
    _write_row(
        packages_sheet,
        1,
        [
            labels["package"],
            labels["value"],
            labels["min_class"],
            labels["certs"],
            labels["state"],
            labels["gap"],
        ],
    )
    rows = list_packages(session, project.id)
    for index, package in enumerate(rows, start=2):
        result = by_package_id.get(str(package.id), {})
        name = package.category.name_en if locale == "en" else package.category.name_az
        _write_row(
            packages_sheet,
            index,
            [
                package.name or name,
                float(package.estimated_value),
                package.min_class.value,
                ", ".join(package.required_certs),
                result.get("state", ""),
                result.get("gap") or "",
            ],
        )

    candidates_sheet = workbook.create_sheet("Candidates")
    _write_row(
        candidates_sheet,
        1,
        [
            labels["package"],
            labels["vendor"],
            labels["score"],
            labels["class"],
            labels["capacity"],
            labels["eligible"],
        ],
    )
    line = 2
    for package in rows:
        result = by_package_id.get(str(package.id), {})
        name = package.category.name_en if locale == "en" else package.category.name_az
        for candidate in result.get("candidates", []):
            _write_row(
                candidates_sheet,
                line,
                [
                    package.name or name,
                    candidate["legal_name"],
                    candidate["total"],
                    candidate["cls"],
                    "fit" if candidate["capacity_fit"] else "small",
                    "yes" if candidate["eligible"] else "no",
                ],
            )
            line += 1

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _write_row(sheet: Worksheet, row: int, values: list[Any]) -> None:
    for column, value in enumerate(values, start=1):
        sheet.cell(row=row, column=column, value=value)


def match_run_payload(row: MatchRunRow) -> dict[str, Any]:
    """``MatchRun`` response body: the stored envelope plus its own identity and timing."""
    return {
        "id": row.id,
        "project_id": row.project_id,
        "ran_at": row.ran_at,
        "params": row.params,
        **row.result,
    }


__all__ = [
    "PACKAGE_AUDIT_FIELDS",
    "PROJECT_AUDIT_FIELDS",
    "ProjectFilters",
    "by_code",
    "create",
    "create_package",
    "delete",
    "delete_package",
    "export_workbook",
    "get",
    "get_package",
    "latest_match",
    "list_packages",
    "list_page",
    "match_run_payload",
    "package_count",
    "patch",
    "patch_package",
    "run_match",
]
