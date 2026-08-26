"""The commission summary — the same content the workbook sheet "5. Nəticə Xülasəsi" carries.

The xlsx export reproduces that sheet's own structure (participant columns, group score
rows, the KO row, the result row, the signature line) directly from
``seed/fixtures/3b699c4f-Rev4_Prekvalifikasiya_TQS2026006.xlsx`` — read with ``openpyxl`` while
building this module, not guessed at — so the workbook a commission chair opens today and the
one this endpoint produces are the same document read two ways. The pdf export carries the
same rows as a flat table; it does not attempt the workbook's cell styling.

Nothing here computes a score. Every number comes from ``application.computed``, already
produced by :mod:`vendoriq_scoring` through :mod:`vendoriq_api.services.evaluation` — this
module only lays it out.
"""

from __future__ import annotations

import uuid
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.orm import Session

from ..errors import ApiError
from ..models import Application, Contact, QualificationCycle, Vendor
from ..models import ScoringModel as ScoringModelRow
from ..models.enums import VendorType
from . import vendors as vendors_service

__all__ = [
    "build_commission_summary_pdf",
    "build_commission_summary_workbook",
    "build_vendor_register_workbook",
]

#: The sheet name to reproduce, verbatim from the Rev4 workbook (Excel's 31-char cap fits it).
SHEET_NAME = "5. Nəticə Xülasəsi"

# Colours lifted from the reference workbook's own cell fills (read with openpyxl), not chosen
# fresh — the point is that the page a manager already knows stays recognisable.
_NAVY = "1F3864"
_BLUE = "2E5DA3"
_ACCENT = "4472C4"
_DARK = "404040"
_GOLD = "C9A227"
_LIGHT = "F2F2F2"
_CREAM = "FFF2CC"
_WHITE = "FFFFFF"

_TITLE = {
    "az": "PREKVALİFİKASİYA NƏTİCƏLƏRİ  —  İCRAİ XÜLASƏ",
    "en": "PREQUALIFICATION RESULTS — EXECUTIVE SUMMARY",
}
_SUBTITLE = {
    "az": "Rəhbərliyin Təsdiqi üçün  |  Tenderə Dəvət Edilməli İştirakçıların Müəyyən edilməsi",
    "en": "For Management Approval | Determining the Participants to Invite to Tender",
}
_LABELS = {
    "indicator": {"az": "Göstərici", "en": "Indicator"},
    "participant": {"az": "İştirakçı", "en": "Participant"},
    "company": {"az": "Şirkətin Adı", "en": "Company name"},
    "reg_year": {"az": "Qeydiyyat ili", "en": "Registration year"},
    "staff": {"az": "Ümumi heyət sayı", "en": "Total staff"},
    "contact": {"az": "Əlaqə şəxsi", "en": "Contact"},
    "phone": {"az": "Telefon", "en": "Phone"},
    "group_head": {"az": "  BAL DETALI (Qrup üzrə)", "en": "  SCORE DETAIL (by group)"},
    "total": {"az": "ÜMUMİ BAL", "en": "TOTAL SCORE"},
    "ko": {"az": "Məcburi Meyar (KO)", "en": "Mandatory criteria (KO)"},
    "ko_pass": {"az": "Keçdi ✓", "en": "Passed ✓"},
    "ko_fail": {"az": "RƏDD ✗", "en": "FAILED ✗"},
    "result": {"az": "NƏTİCƏ / TÖVSİYƏ", "en": "RESULT / RECOMMENDATION"},
    "notes_head": {
        "az": "  KOMİSSİYANIN QEYDLƏRİ VƏ ƏSASLANDIRMA",
        "en": "  COMMISSION NOTES AND RATIONALE",
    },
    "chair": {"az": "Komissiya Sədri / Tarix:", "en": "Commission Chair / Date:"},
    "management": {"az": "Rəhbərin Təsdiqi / Tarix:", "en": "Management Approval / Date:"},
    "ko_reject": {"az": "KO — RƏDD", "en": "KO — REJECT"},
}


def _label(key: str, locale: str) -> str:
    return _LABELS[key].get(locale, _LABELS[key]["en"])


@dataclass(frozen=True, slots=True)
class _SummaryRow:
    vendor: Vendor
    application: Application
    contact: Contact | None


def _rows(session: Session, cycle: QualificationCycle) -> list[_SummaryRow]:
    """One row per application in the cycle, alphabetical by legal name.

    The Rev4 workbook's own participant order is the intake order and is not preserved by
    any stored field, so this is the closest deterministic ordering the data supports —
    documented as a deviation in the final report.
    """
    applications = session.scalars(
        select(Application).where(Application.cycle_id == cycle.id)
    ).all()
    rows: list[_SummaryRow] = []
    for application in applications:
        vendor = session.get(Vendor, application.vendor_id)
        if vendor is None:  # pragma: no cover - FK guarantees this
            continue
        contact = vendors_service.primary_contact(session, vendor.id)
        rows.append(_SummaryRow(vendor=vendor, application=application, contact=contact))
    rows.sort(key=lambda row: row.vendor.legal_name.lower())
    return rows


def _model_row_or_404(session: Session, version: str) -> ScoringModelRow:
    row = session.get(ScoringModelRow, version)
    if row is None:
        raise ApiError(404, "not_found", f"No such scoring model version {version!r}.")
    return row


def _decision_text(computed: dict[str, Any], model_row: ScoringModelRow, locale: str) -> str:
    ko = bool(computed.get("ko"))
    cls = computed.get("cls")
    if not ko or not cls or cls == "KO":
        return _label("ko_reject", locale)
    band = next((b for b in model_row.classes if b["cls"] == cls), None)
    label = band.get(f"label_{locale}", cls) if band else cls
    return f"{cls} — {label}"


def _explanation(model_row: ScoringModelRow, locale: str) -> str:
    thresholds = ", ".join(f"{band['cls']} ({int(band['min'])}+)" for band in model_row.classes)
    if locale == "az":
        return (
            f"Sinif hədləri (100 baldan): {thresholds}. Məcburi meyarlardan (KO) birini "
            "təmin etməyən iştirakçı ümumi baldan asılı olmayaraq avtomatik RƏDD alır."
        )
    return (
        f"Class thresholds (out of 100): {thresholds}. A participant failing any mandatory "
        "(KO) criterion is automatically REJECTED regardless of the total score."
    )


def _staff_count(row: _SummaryRow, model_row: ScoringModelRow) -> Any:
    """``E.1`` is permanent staff on the subcontractor model only (sup-1's E group is
    commercial terms, not headcount) — a supplier cycle shows the column empty rather than a
    number that means something else."""
    if model_row.vendor_type is not VendorType.SUB:
        return None
    raw = row.application.raw_snapshot or {}
    return raw.get("E.1")


# ── xlsx ─────────────────────────────────────────────────────────────────────
def build_commission_summary_workbook(
    session: Session, cycle: QualificationCycle, *, locale: str = "az"
) -> Workbook:
    """Reproduce the layout of sheet "5. Nəticə Xülasəsi" for this cycle's applications."""
    model_row = _model_row_or_404(session, cycle.scoring_model_version)
    rows = _rows(session, cycle)
    groups = list(model_row.groups)

    first_col = 3  # column C — column B carries the row label, mirroring the workbook
    last_col = first_col + max(len(rows), 1) - 1
    pad_col = last_col + 1

    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = SHEET_NAME[:31]
    sheet.sheet_view.showGridLines = False

    sheet.column_dimensions["A"].width = 2
    sheet.column_dimensions["B"].width = 28
    for col in range(first_col, last_col + 1):
        sheet.column_dimensions[get_column_letter(col)].width = 14
    sheet.column_dimensions[get_column_letter(pad_col)].width = 2

    def band(
        row: int,
        text: str,
        *,
        fill: str,
        size: float = 10,
        bold: bool = True,
        color: str = _WHITE,
        align: str = "left",
        end_col: int = pad_col,
    ) -> None:
        sheet.merge_cells(start_row=row, start_column=2, end_row=row, end_column=end_col)
        cell = sheet.cell(row=row, column=2, value=text)
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.font = Font(name="Calibri", size=size, bold=bold, color=color)
        cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)

    def row_label(row: int, text: str, *, fill: str = _LIGHT, color: str = _DARK) -> None:
        cell = sheet.cell(row=row, column=2, value=text)
        cell.font = Font(name="Calibri", size=9, bold=True, color=color)
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    def row_values(
        row: int,
        values: list[Any],
        *,
        fill: str | None = None,
        bold: bool = False,
        color: str = "000000",
    ) -> None:
        for index, value in enumerate(values):
            cell = sheet.cell(row=row, column=first_col + index, value=value)
            cell.font = Font(name="Calibri", size=9, bold=bold, color=color)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if fill is not None:
                cell.fill = PatternFill("solid", fgColor=fill)

    band(2, _TITLE.get(locale, _TITLE["en"]), fill=_NAVY, size=14)
    sheet.row_dimensions[2].height = 30
    subtitle = f"{cycle.name} · {_SUBTITLE.get(locale, _SUBTITLE['en'])}"
    band(3, subtitle, fill=_BLUE, size=9, bold=False)
    sheet.row_dimensions[3].height = 18

    header_cell = sheet.cell(row=5, column=2, value=_label("indicator", locale))
    header_cell.font = Font(name="Calibri", size=9, bold=True, color=_WHITE)
    header_cell.fill = PatternFill("solid", fgColor=_ACCENT)
    header_cell.alignment = Alignment(horizontal="center", vertical="center")
    participant = _label("participant", locale)
    for index in range(len(rows)):
        cell = sheet.cell(row=5, column=first_col + index, value=f"{participant} №{index + 1}")
        cell.font = Font(name="Calibri", size=8, bold=True, color=_WHITE)
        cell.fill = PatternFill("solid", fgColor=_ACCENT)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.row_dimensions[5].height = 22

    row_label(6, _label("company", locale))
    row_label(7, _label("reg_year", locale))
    row_label(8, _label("staff", locale))
    row_label(9, _label("contact", locale))
    row_label(10, _label("phone", locale))
    row_values(6, [r.vendor.legal_name for r in rows])
    row_values(7, [r.vendor.registration_year or "" for r in rows])
    row_values(8, [_staff_count(r, model_row) or "" for r in rows])
    row_values(9, [r.contact.name if r.contact else "" for r in rows])
    row_values(10, [r.contact.phone if r.contact and r.contact.phone else "" for r in rows])

    band(12, _label("group_head", locale), fill=_ACCENT, size=10)

    detail_row = 13
    for group in groups:
        name = group.get(f"name_{locale}", group.get("name_en", group["group"]))
        row_label(detail_row, f"{group['group']}. {name} (/{int(group['max'])})", fill=_WHITE)
        group_points = [
            (r.application.computed or {}).get("groups", {}).get(group["group"], "") for r in rows
        ]
        row_values(detail_row, group_points)
        detail_row += 1

    total_row = detail_row + 1
    total_label = f"{_label('total', locale)} ({int(model_row_total_max(model_row))})"
    row_label(total_row, total_label, fill=_NAVY, color=_WHITE)
    total_label_font = Font(name="Calibri", size=11, bold=True, color=_WHITE)
    sheet.cell(row=total_row, column=2).font = total_label_font
    row_values(
        total_row,
        [(r.application.computed or {}).get("total", "") for r in rows],
        fill=_NAVY,
        bold=True,
        color=_WHITE,
    )
    for index in range(len(rows)):
        sheet.cell(row=total_row, column=first_col + index).font = Font(
            name="Calibri", size=11, bold=True, color=_WHITE
        )

    ko_row = total_row + 1
    row_label(ko_row, _label("ko", locale), fill=_DARK, color=_WHITE)
    row_values(
        ko_row,
        [
            _label("ko_pass", locale)
            if bool((r.application.computed or {}).get("ko"))
            else _label("ko_fail", locale)
            for r in rows
        ],
        fill=_DARK,
        bold=True,
        color=_WHITE,
    )

    result_row = ko_row + 1
    row_label(result_row, _label("result", locale), fill=_GOLD, color=_WHITE)
    row_values(
        result_row,
        [_decision_text(r.application.computed or {}, model_row, locale) for r in rows],
        fill=_GOLD,
        bold=True,
        color=_WHITE,
    )
    sheet.row_dimensions[total_row].height = 22
    sheet.row_dimensions[result_row].height = 22

    notes_row = result_row + 2
    band(notes_row, _label("notes_head", locale), fill=_BLUE, size=10)

    explain_row = notes_row + 1
    sheet.merge_cells(
        start_row=explain_row, start_column=2, end_row=explain_row + 3, end_column=pad_col
    )
    explain_cell = sheet.cell(row=explain_row, column=2, value=_explanation(model_row, locale))
    explain_cell.fill = PatternFill("solid", fgColor=_CREAM)
    explain_cell.font = Font(name="Calibri", size=10, color=_DARK)
    explain_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    sig_row = explain_row + 5
    midpoint = first_col + max(len(rows), 1) // 2 - 1
    for start, end, key in ((2, midpoint, "chair"), (midpoint + 1, pad_col, "management")):
        sheet.merge_cells(start_row=sig_row, start_column=start, end_row=sig_row, end_column=end)
        cell = sheet.cell(row=sig_row, column=start, value=_label(key, locale))
        cell.font = Font(name="Calibri", size=10, bold=True, color=_DARK)
        cell.fill = PatternFill("solid", fgColor=_LIGHT)
        cell.alignment = Alignment(horizontal="left", vertical="center")
    sheet.row_dimensions[sig_row].height = 20

    return workbook


def model_row_total_max(model_row: ScoringModelRow) -> float:
    """Sum of the group maxima — 100 for both shipped models (ADR-014: not a stored column)."""
    return sum(float(group["max"]) for group in model_row.groups)


# ── pdf ──────────────────────────────────────────────────────────────────────
#: Debian/Ubuntu's ``fonts-dejavu-core`` package path (``infra/Dockerfile.api`` installs it).
#: Azerbaijani letters absent from the PDF base-14 fonts' encoding (ə, ş, ç, ğ, ı) need a
#: Unicode TTF; without the package the export still produces a valid PDF, just in Helvetica
#: with those letters substituted by the font's ``.notdef`` glyph.
_DEJAVU_PATHS = (
    Path("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"),
    Path("/usr/share/fonts/dejavu/DejaVuSans.ttf"),
)
_DEJAVU_BOLD_NAME = "DejaVuSans-Bold.ttf"


def _find_unicode_font() -> Path | None:
    for path in _DEJAVU_PATHS:
        if path.is_file():
            return path
    return None


def build_commission_summary_pdf(
    session: Session, cycle: QualificationCycle, *, locale: str = "az"
) -> bytes:
    """The same rows as the xlsx export, as a flat table — content, not the cell styling."""
    from fpdf import FPDF

    model_row = _model_row_or_404(session, cycle.scoring_model_version)
    rows = _rows(session, cycle)
    groups = list(model_row.groups)

    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=12)
    pdf.add_page()

    font_path = _find_unicode_font()
    if font_path is not None:
        pdf.add_font("Body", "", str(font_path))
        bold_path = font_path.with_name(_DEJAVU_BOLD_NAME)
        pdf.add_font("Body", "B", str(bold_path) if bold_path.is_file() else str(font_path))
        family = "Body"
    else:  # pragma: no cover - exercised only when the font package is missing
        family = "Helvetica"

    label_w = 55.0
    usable = pdf.w - pdf.l_margin - pdf.r_margin - label_w
    col_w = max(18.0, usable / max(len(rows), 1))

    def hex_rgb(value: str) -> tuple[int, int, int]:
        return int(value[0:2], 16), int(value[2:4], 16), int(value[4:6], 16)

    def band_row(text: str, *, fill: str, size: float = 12) -> None:
        pdf.set_font(family, "B", size)
        pdf.set_text_color(255, 255, 255)
        pdf.set_fill_color(*hex_rgb(fill))
        pdf.cell(0, 9, text, align="C", fill=True, new_x="LMARGIN", new_y="NEXT")

    def data_row(
        label: str,
        values: list[Any],
        *,
        fill: str | None = None,
        bold: bool = False,
        text_color: tuple[int, int, int] = (0, 0, 0),
    ) -> None:
        pdf.set_font(family, "B" if bold else "", 8)
        pdf.set_text_color(*text_color)
        if fill is not None:
            pdf.set_fill_color(*hex_rgb(fill))
        pdf.cell(label_w, 6.5, label, border=1, fill=fill is not None)
        for value in values:
            text = "" if value is None else str(value)
            pdf.cell(col_w, 6.5, text, border=1, align="C", fill=fill is not None)
        pdf.ln(6.5)

    band_row(_TITLE.get(locale, _TITLE["en"]), fill=_NAVY, size=13)
    band_row(f"{cycle.name} · {_SUBTITLE.get(locale, _SUBTITLE['en'])}", fill=_BLUE, size=9)
    pdf.ln(2)

    pdf.set_text_color(255, 255, 255)
    pdf.set_fill_color(*hex_rgb(_ACCENT))
    pdf.set_font(family, "B", 8)
    pdf.cell(label_w, 7, _label("indicator", locale), border=1, fill=True)
    for index in range(len(rows)):
        pdf.cell(col_w, 7, f"№{index + 1}", border=1, align="C", fill=True)
    pdf.ln(7)
    pdf.set_text_color(0, 0, 0)

    contact_names = [r.contact.name if r.contact else "" for r in rows]
    contact_phones = [r.contact.phone if r.contact and r.contact.phone else "" for r in rows]
    data_row(_label("company", locale), [r.vendor.legal_name for r in rows])
    data_row(_label("reg_year", locale), [r.vendor.registration_year or "" for r in rows])
    data_row(_label("staff", locale), [_staff_count(r, model_row) or "" for r in rows])
    data_row(_label("contact", locale), contact_names)
    data_row(_label("phone", locale), contact_phones)

    pdf.ln(1)
    band_row(_label("group_head", locale).strip(), fill=_ACCENT, size=10)
    for group in groups:
        name = group.get(f"name_{locale}", group.get("name_en", group["group"]))
        group_points = [
            (r.application.computed or {}).get("groups", {}).get(group["group"], "") for r in rows
        ]
        data_row(f"{group['group']}. {name} (/{int(group['max'])})", group_points)

    data_row(
        f"{_label('total', locale)} ({int(model_row_total_max(model_row))})",
        [(r.application.computed or {}).get("total", "") for r in rows],
        fill=_NAVY,
        bold=True,
        text_color=(255, 255, 255),
    )
    data_row(
        _label("ko", locale),
        [
            _label("ko_pass", locale)
            if bool((r.application.computed or {}).get("ko"))
            else _label("ko_fail", locale)
            for r in rows
        ],
        fill=_DARK,
        bold=True,
        text_color=(255, 255, 255),
    )
    data_row(
        _label("result", locale),
        [_decision_text(r.application.computed or {}, model_row, locale) for r in rows],
        fill=_GOLD,
        bold=True,
        text_color=(255, 255, 255),
    )

    pdf.ln(4)
    pdf.set_font(family, "B", 10)
    pdf.set_text_color(0, 0, 0)
    pdf.multi_cell(0, 6, _label("notes_head", locale).strip(), new_x="LMARGIN", new_y="NEXT")
    pdf.set_font(family, "", 9)
    pdf.multi_cell(0, 5, _explanation(model_row, locale), new_x="LMARGIN", new_y="NEXT")

    pdf.ln(8)
    pdf.set_font(family, "B", 9)
    half = pdf.w / 2
    pdf.cell(half - pdf.l_margin, 6, _label("chair", locale))
    pdf.cell(half - pdf.r_margin, 6, _label("management", locale), new_x="LMARGIN", new_y="NEXT")

    return bytes(pdf.output())


# ── vendor register (screen 16) ─────────────────────────────────────────────
#: A generous cap, not a real page — `exportVendors` follows the current filter, never the
#: whole register (contract note), and spec §13's own performance target (1 000 vendors under
#: one second) is well inside it.
_REGISTER_EXPORT_LIMIT = 10_000

_REGISTER_LABELS: dict[str, dict[str, str]] = {
    "sheet": {"az": "Vendor reyestri", "en": "Vendor register"},
    "name": {"az": "Hüquqi ad", "en": "Legal name"},
    "voen": {"az": "VÖEN", "en": "Tax ID"},
    "type": {"az": "Tip", "en": "Type"},
    "score": {"az": "Bal", "en": "Score"},
    "class": {"az": "Sinif", "en": "Class"},
    "status": {"az": "Status", "en": "Status"},
    "region": {"az": "Region", "en": "Region"},
    "source": {"az": "Mənbə", "en": "Source"},
    "updated": {"az": "Yenilənib", "en": "Updated"},
}

_TYPE_LABEL: dict[str, dict[str, str]] = {
    "sub": {"az": "Subpodratçı", "en": "Subcontractor"},
    "sup": {"az": "Tədarükçü", "en": "Supplier"},
    "both": {"az": "Hər ikisi", "en": "Both"},
}

_STATUS_LABEL: dict[str, dict[str, str]] = {
    "registered": {"az": "Qeydiyyatdan keçib", "en": "Registered"},
    "invited": {"az": "Dəvət edilib", "en": "Invited"},
    "in_progress": {"az": "Doldurulur", "en": "In progress"},
    "submitted": {"az": "Təqdim edilib", "en": "Submitted"},
    "under_review": {"az": "Baxılır", "en": "Under review"},
    "information_requested": {"az": "Əlavə məlumat gözlənilir", "en": "Information requested"},
    "prequalified": {"az": "Prekvalifikasiya keçib", "en": "Prequalified"},
    "rejected": {"az": "Rədd edilib", "en": "Rejected"},
    "suspended": {"az": "Dayandırılıb", "en": "Suspended"},
}

_SOURCE_LABEL: dict[str, dict[str, str]] = {
    "registry": {"az": "Reyestr", "en": "Registry"},
    "api": {"az": "ERP API", "en": "ERP API"},
    "document": {"az": "Sənəd", "en": "Document"},
    "portal": {"az": "Portal", "en": "Portal"},
    "excel": {"az": "Excel forma", "en": "Excel form"},
    "manual": {"az": "Əl ilə", "en": "Manual"},
}


def _localised(table: dict[str, dict[str, str]], key: str | None, locale: str) -> str:
    if key is None:
        return ""
    entry = table.get(key)
    if entry is None:
        return key
    return entry.get(locale, entry.get("en", key))


def build_vendor_register_workbook(
    session: Session,
    filters: vendors_service.VendorFilters,
    *,
    principal_vendor_id: uuid.UUID | None = None,
    locale: str = "az",
) -> Workbook:
    """The register as it stands under the caller's current filter (contract note on
    ``exportVendors``: "takes the same filters as ``listVendors``... follows the current
    filter, not the whole register"). Every row is a plain read of the vendor and its
    already-computed latest result — no scoring happens here."""
    rows, _total = vendors_service.list_page(
        session,
        filters,
        page=1,
        page_size=_REGISTER_EXPORT_LIMIT,
        principal_vendor_id=principal_vendor_id,
    )
    if filters.classes:
        wanted = set(filters.classes)
        rows = [v for v in rows if vendors_service.latest_result(session, v.id).cls in wanted]

    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = _localised(_REGISTER_LABELS, "sheet", locale)[:31]
    sheet.freeze_panes = "A2"

    columns = ("name", "voen", "type", "score", "class", "status", "region", "source", "updated")
    header_font = Font(name="Calibri", size=10, bold=True, color=_WHITE)
    header_fill = PatternFill("solid", fgColor=_ACCENT)
    for col, key in enumerate(columns, start=1):
        cell = sheet.cell(row=1, column=col, value=_localised(_REGISTER_LABELS, key, locale))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")
        sheet.column_dimensions[get_column_letter(col)].width = 20
    sheet.column_dimensions["A"].width = 32
    sheet.row_dimensions[1].height = 20

    for row_index, vendor in enumerate(rows, start=2):
        result = vendors_service.latest_result(session, vendor.id)
        source = vendors_service.primary_source(session, vendor.id)
        values = (
            vendor.legal_name,
            vendor.voen or "",
            _localised(_TYPE_LABEL, vendor.type.value, locale),
            result.total if result.total is not None else "",
            result.cls.value if result.cls is not None else "",
            _localised(_STATUS_LABEL, vendor.status.value, locale),
            vendor.region or "",
            _localised(_SOURCE_LABEL, source.value if source is not None else None, locale),
            vendor.updated_at.date().isoformat(),
        )
        for col, value in enumerate(values, start=1):
            sheet.cell(row=row_index, column=col, value=value)

    return workbook
