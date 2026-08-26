"""``python -m vendoriq_excel_import parse <file> [--json]``.

The importer runs before anything else exists — no database, no API, no browser — so it
needs a way to answer "what would this file import as?" from a terminal. ``--json`` prints
the whole parse for piping; without it you get the summary an officer would read first.
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path
from typing import Any

from .form import parse_application_form
from .workbook import parse_scoring_workbook, to_seed_rows

#: Sheet titles that only a scoring workbook has.
_WORKBOOK_MARKERS = ("Bal Hesablaması", "Bal Sxemi")


def detect_kind(path: Path) -> str:
    """``"workbook"`` or ``"form"``, decided by the sheet names rather than the file name."""
    from openpyxl import load_workbook

    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        titles = " | ".join(workbook.sheetnames)
    finally:
        workbook.close()
    return "workbook" if any(marker in titles for marker in _WORKBOOK_MARKERS) else "form"


def _parse(path: Path, kind: str) -> dict[str, Any]:
    if kind == "workbook":
        vendors = parse_scoring_workbook(path)
        return {
            "kind": "scoring_workbook",
            "source_file": path.name,
            "vendors": [vendor.as_dict() for vendor in vendors],
            "seed_rows": to_seed_rows(vendors),
        }
    application = parse_application_form(path)
    return {"kind": "application_form", **application.as_dict()}


def _print_summary(parsed: dict[str, Any]) -> None:
    if parsed["kind"] == "scoring_workbook":
        print(f"{parsed['source_file']}: scoring workbook, {len(parsed['vendors'])} participants")
        for vendor in parsed["vendors"]:
            name = str(vendor["name"]).strip()
            print(f"  {name:<28} {vendor['sheet_total']!s:>6}  {vendor['sheet_decision']}")
        return

    print(f"{parsed['source_file']}: application form")
    vendor = parsed["vendor"]
    print(f"  vendor    {vendor.get('name') or '—'}  (VÖEN {vendor.get('voen') or '—'})")
    meta = parsed["meta"]
    print(f"  cycle     {meta.get('project_name') or '—'} / TQS-{meta.get('project_code') or '—'}")
    print(f"  answers   {len(parsed['answers'])}")
    for code, rows in parsed["tables"].items():
        print(f"  table     {code}: {len(rows)} rows")
    uploaded = sum(1 for status in parsed["documents"].values() if status == "uploaded")
    print(f"  documents {uploaded} of {len(parsed['documents'])} uploaded")
    print(f"  warnings  {len(parsed['warnings'])}")
    for warning in parsed["warnings"]:
        print(f"    [{warning['severity']}] {warning['code']}: {warning['message_en']}")


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(prog="python -m vendoriq_excel_import")
    subcommands = parser.add_subparsers(dest="command", required=True)
    parse_cmd = subcommands.add_parser(
        "parse", help="parse an application form or scoring workbook"
    )
    parse_cmd.add_argument("file", type=Path)
    parse_cmd.add_argument("--json", action="store_true", help="print the full parse as JSON")
    parse_cmd.add_argument(
        "--kind",
        choices=("auto", "form", "workbook"),
        default="auto",
        help="override the automatic form/workbook detection",
    )
    args = parser.parse_args(argv)

    path: Path = args.file
    if not path.is_file():
        print(f"no such file: {path}", file=sys.stderr)
        return 2

    kind = detect_kind(path) if args.kind == "auto" else str(args.kind)
    parsed = _parse(path, kind)
    if args.json:
        json.dump(parsed, sys.stdout, ensure_ascii=False, indent=2, sort_keys=False)
        sys.stdout.write("\n")
    else:
        _print_summary(parsed)
    return 0


if __name__ == "__main__":  # pragma: no cover - exercised through the CLI test
    raise SystemExit(main())
