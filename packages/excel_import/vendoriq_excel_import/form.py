"""Reading the 11-sheet vendor application form.

Answers are found by the **code in column B** (``A.1`` … ``G.7``), never by row number, so
a sheet that grew a row still parses (spec §6.1). Section tables become lists of dicts, the
checklist sheet becomes ``{document code: status}``, and the cover sheet supplies the cycle
metadata and the vendor's identity.

Nothing here reads the clock. The "is this certificate stale?" question is answered against
the date the form itself carries ("Forma Göndərilmə Tarixi"), which is both the honest
reference — the certificate had to be fresh when the vendor applied, not when an officer
happened to open the file — and the reason the fixtures parse identically forever.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from .catalog import (
    CODE_COL,
    COVER_LABELS,
    COVER_SHEET_INDEX,
    COVER_VENDOR_LABELS,
    DOCUMENT_CATALOG,
    DOCUMENT_COMPLETION_LABEL,
    DOCUMENT_STATUS_WORDS,
    DOCUMENTS_SHEET_INDEX,
    FIELD_CATALOG,
    MANDATORY_FIELD_CODES,
    SECTION_SHEETS,
    TABLE_DEFS,
    UNIT_COL,
    VENDOR_FIELD_CODES,
    FieldDef,
    TableDef,
)
from .derive import derive_indicators
from .normalise import (
    ImportWarning,
    Warnings,
    clean_text,
    is_blank,
    months_between,
    normalise_bool,
    normalise_date,
    normalise_number,
    normalise_percent,
    parse_iso,
    percent_style,
    split_multi_value,
)
from .sheets import (
    code_row_index,
    label_row_index,
    normalise_label,
    sheet_by_index,
)

#: How long a tax-clearance certificate (document A-05) stays fresh — brief §1.5.
TAX_CLEARANCE_VALID_MONTHS = 3

#: Unit hints the sheets print in column D, mapped to the unit an observation stores.
#: Anything else in that column is a format hint ("Mətn", "dd.mm.yyyy"), not a unit.
_UNITS: dict[str, str] = {
    "azn": "AZN",
    "m²": "m2",
    "nəfər": "person",
    "ədəd": "unit",
    "saat": "hour",
    "şirkət": "company",
    # Azerbaijani "İl" casefolds to "i" + a combining dot, not to "il", so both spellings
    # of the year hint have to be listed.
    "il": "year",
    "i̇l": "year",
    "katsayı": "ratio",
}


@dataclass(slots=True)
class ParsedApplication:
    """Everything one application form says, normalised.

    ``answers`` is keyed by field code and holds normalised scalars (ISO date strings,
    booleans, floats, text, or a list where one cell held two values). ``tables`` holds the
    three section tables. ``documents`` maps a checklist code to its status.
    """

    source_file: str
    vendor: dict[str, Any] = field(default_factory=dict)
    answers: dict[str, Any] = field(default_factory=dict)
    tables: dict[str, list[dict[str, Any]]] = field(default_factory=dict)
    documents: dict[str, str] = field(default_factory=dict)
    meta: dict[str, Any] = field(default_factory=dict)
    #: The checklist sheet in full — name, mandatory flag, status and the vendor's note.
    document_details: list[dict[str, Any]] = field(default_factory=list)
    #: Raw indicators for the scoring engine, derived from the answers (see ``derive.py``).
    derived: dict[str, float | None] = field(default_factory=dict)
    #: Field code -> unit, taken from the sheet's own unit column.
    units: dict[str, str | None] = field(default_factory=dict)
    warnings: list[ImportWarning] = field(default_factory=list)

    def to_observations(
        self,
        source: str = "excel",
        source_ref: str | None = None,
    ) -> list[dict[str, Any]]:
        """One ``field_observation`` row per answer and per table (ADR-004).

        Values are wrapped as ``{"value": …}`` so a table and a number live in the same
        JSONB column. The caller supplies ``vendor_id``, ``observed_at`` and ``entered_by``;
        the importer knows none of those.
        """
        ref = source_ref if source_ref is not None else self.source_file
        rows: list[dict[str, Any]] = []
        for code, value in self.answers.items():
            rows.append(
                {
                    "field_code": code,
                    "value": {"value": value},
                    "unit": self.units.get(code),
                    "source": source,
                    "source_ref": ref,
                }
            )
        for code, table in self.tables.items():
            rows.append(
                {
                    "field_code": code,
                    "value": {"value": table},
                    "unit": None,
                    "source": source,
                    "source_ref": ref,
                }
            )
        return rows

    def as_dict(self) -> dict[str, Any]:
        """A JSON-serialisable view — what the CLI prints and the fixtures compare."""
        return {
            "source_file": self.source_file,
            "vendor": self.vendor,
            "meta": self.meta,
            "answers": self.answers,
            "tables": self.tables,
            "documents": self.documents,
            "document_details": self.document_details,
            "derived": self.derived,
            "units": self.units,
            "warnings": [w.as_dict() for w in self.warnings],
        }


def parse_application_form(path: Path | str) -> ParsedApplication:
    """Read the 11-sheet vendor application form at ``path``."""
    source = Path(path)
    workbook = load_workbook(source, data_only=True, read_only=False)
    try:
        warnings = Warnings()
        result = ParsedApplication(source_file=source.name)

        _parse_cover(workbook, result, warnings)
        for spec in SECTION_SHEETS:
            sheet = sheet_by_index(workbook, spec.index)
            if sheet is None:
                warnings.add(
                    "missing_sheet",
                    message_en=f"Section {spec.section} sheet is missing from the workbook.",
                    message_az=f"{spec.section} bölməsinin vərəqəsi faylda yoxdur.",
                    severity="error",
                    field_code=spec.section,
                )
                continue
            _parse_section(sheet, spec.section, spec.answer_col, result, warnings)
            _check_currency_label(sheet, spec.section, warnings)
        for table_def in TABLE_DEFS:
            sheet = sheet_by_index(workbook, table_def.sheet_index)
            if sheet is not None:
                result.tables[table_def.code] = _parse_table(sheet, table_def, warnings)
        _parse_documents(workbook, result, warnings)

        _fill_vendor(result)
        _check_mandatory(result, warnings)
        _check_tax_clearance(result, warnings)
        result.derived = derive_indicators(result.answers, result.tables, result.meta)
        result.warnings = warnings.items
        return result
    finally:
        workbook.close()


# --------------------------------------------------------------------------------------
# Cover sheet
# --------------------------------------------------------------------------------------


def _cover_value(raw: object) -> object:
    """The cover sheet fills its vendor block by formula: unfilled cells read ``0``.

    Bracketed instructions ("(dolduruldukdan sonra avtomatik)") are placeholders, not data.
    """
    if raw == 0:
        return None
    if isinstance(raw, str):
        text = raw.strip()
        if text.startswith("(") and text.endswith(")"):
            return None
    return raw


def _parse_cover(workbook: Any, result: ParsedApplication, warnings: Warnings) -> None:
    sheet = sheet_by_index(workbook, COVER_SHEET_INDEX)
    if sheet is None:
        warnings.add(
            "missing_sheet",
            message_en="Cover sheet is missing from the workbook.",
            message_az="Üz səhifə vərəqəsi faylda yoxdur.",
            severity="error",
        )
        return
    labels = label_row_index(sheet, CODE_COL)
    for label, key in COVER_LABELS.items():
        row = labels.get(normalise_label(label))
        if row is None:
            continue
        raw = _cover_value(sheet.cell(row=row, column=3).value)
        if key in {"issued_on", "due_on"}:
            iso, status = normalise_date(raw)
            result.meta[key] = iso
            if status == "unparsable" and not is_blank(raw):
                warnings.add(
                    "unparsable_date",
                    message_en=f"Cover sheet: {label} is not a date.",
                    message_az=f"Üz səhifə: «{label}» tarix formatında deyil.",
                    sheet=sheet.title,
                    cell=f"C{row}",
                    raw_value=raw,
                )
        else:
            result.meta[key] = clean_text(raw)
    result.meta["source_file"] = result.source_file

    for label, key in COVER_VENDOR_LABELS.items():
        row = labels.get(normalise_label(label))
        if row is None:
            continue
        raw = _cover_value(sheet.cell(row=row, column=3).value)
        value = clean_text(raw)
        if value is not None:
            result.vendor.setdefault(key, value)


# --------------------------------------------------------------------------------------
# Section sheets
# --------------------------------------------------------------------------------------


def _unit_for(sheet: Worksheet, row: int) -> str | None:
    hint = clean_text(sheet[f"{UNIT_COL}{row}"].value)
    if hint is None:
        return None
    return _UNITS.get(hint.casefold())


def _parse_section(
    sheet: Worksheet,
    section: str,
    answer_col: str,
    result: ParsedApplication,
    warnings: Warnings,
) -> None:
    rows = code_row_index(sheet, CODE_COL)
    for code, definition in FIELD_CATALOG.items():
        if definition.section != section or definition.kind in {"table", "calc"}:
            continue
        row = rows.get(code)
        if row is None:
            continue
        cell = f"{answer_col}{row}"
        raw = sheet[cell].value
        if is_blank(raw):
            continue
        result.units[code] = _unit_for(sheet, row)
        # A filled cell always produces a key, even when it normalises to ``None``:
        # "Müddətsiz" in an expiry field is an answer ("no expiry"), not a blank.
        result.answers[code] = _normalise_answer(definition, raw, sheet.title, cell, warnings)


def _normalise_answer(
    definition: FieldDef,
    raw: object,
    sheet_title: str,
    cell: str,
    warnings: Warnings,
) -> Any:
    """One answer cell -> the value the system stores, reporting whatever looked wrong."""
    code = definition.code
    shown = clean_text(raw)

    if definition.kind == "date":
        iso, status = normalise_date(raw)
        if status == "no_expiry":
            warnings.add(
                "no_expiry_literal",
                message_en=(f"{code}: «{shown}» means no expiry date — stored as open-ended."),
                message_az=(f"{code}: «{shown}» — bitmə tarixi yoxdur, müddətsiz kimi saxlanıldı."),
                severity="info",
                field_code=code,
                sheet=sheet_title,
                cell=cell,
                raw_value=raw,
            )
            return None
        if status == "unparsable":
            warnings.add(
                "unparsable_date",
                message_en=f"{code}: «{shown}» is not a date; kept as text.",
                message_az=f"{code}: «{shown}» tarix deyil; mətn kimi saxlanıldı.",
                field_code=code,
                sheet=sheet_title,
                cell=cell,
                raw_value=raw,
            )
            return clean_text(raw)
        return iso

    if definition.kind == "bool":
        parsed = normalise_bool(raw)
        if parsed is None:
            warnings.add(
                "unparsable_value",
                message_en=(f"{code}: «{shown}» is not a Var/Yoxdur answer; kept as text."),
                message_az=(f"{code}: «{shown}» Var/Yoxdur cavabı deyil; mətn kimi saxlanıldı."),
                field_code=code,
                sheet=sheet_title,
                cell=cell,
                raw_value=raw,
            )
            return clean_text(raw)
        return parsed

    parts = split_multi_value(raw)
    if parts is not None:
        joined = ", ".join(parts)
        warnings.add(
            "multi_value_cell",
            message_en=(
                f"{code}: the cell holds {len(parts)} values ({joined}); stored as a list."
            ),
            message_az=(
                f"{code}: xanada {len(parts)} dəyər var ({joined}); siyahı kimi saxlanıldı."
            ),
            field_code=code,
            sheet=sheet_title,
            cell=cell,
            raw_value=raw,
        )
        if definition.kind == "number":
            numbers = [normalise_number(part) for part in parts]
            return [n for n in numbers if n is not None]
        return parts

    if definition.kind == "number":
        number = normalise_number(raw)
        if number is None:
            warnings.add(
                "unparsable_value",
                message_en=f"{code}: «{shown}» is not a number; kept as text.",
                message_az=f"{code}: «{shown}» rəqəm deyil; mətn kimi saxlanıldı.",
                field_code=code,
                sheet=sheet_title,
                cell=cell,
                raw_value=raw,
            )
            return clean_text(raw)
        return number

    return clean_text(raw)


def _check_currency_label(sheet: Worksheet, section: str, warnings: Warnings) -> None:
    """The B sheet's subtitle says USD while every unit cell under it says AZN (ADR-007)."""
    if section != "B":
        return
    for row in range(1, 6):
        text = clean_text(sheet[f"{CODE_COL}{row}"].value)
        if text and "USD" in text:
            warnings.add(
                "currency_label_mismatch",
                message_en=(
                    f"Section B header says «{text}» but the figures are AZN; stored as AZN."
                ),
                message_az=(
                    f"B bölməsinin başlığı «{text}» yazır, "
                    f"lakin rəqəmlər AZN-dir; AZN kimi saxlanıldı."
                ),
                severity="info",
                field_code="B",
                sheet=sheet.title,
                cell=f"{CODE_COL}{row}",
                raw_value=text,
            )
            return


# --------------------------------------------------------------------------------------
# Section tables
# --------------------------------------------------------------------------------------


def _parse_table(sheet: Worksheet, table: TableDef, warnings: Warnings) -> list[dict[str, Any]]:
    header = _table_header_row(sheet, table)
    if header is None:
        return []
    number_col = sheet[f"{CODE_COL}1"].column
    rows: list[dict[str, Any]] = []
    percent_styles: set[str] = set()

    for row in range(header + 1, sheet.max_row + 1):
        marker = sheet.cell(row=row, column=number_col).value
        if not isinstance(marker, int):
            break
        entry: dict[str, Any] = {}
        for column in table.columns:
            cell = sheet.cell(row=row, column=number_col + column.offset)
            raw = cell.value
            if is_blank(raw):
                entry[column.key] = None
                continue
            if column.kind == "percent":
                style = percent_style(raw)
                if style is not None:
                    percent_styles.add(style)
                entry[column.key] = normalise_percent(raw)
            elif column.kind == "number":
                entry[column.key] = normalise_number(raw)
            elif column.kind == "date":
                iso, status = normalise_date(raw)
                entry[column.key] = iso if status == "ok" and iso else clean_text(raw)
            else:
                entry[column.key] = clean_text(raw)
        if any(value is not None for value in entry.values()):
            rows.append(entry)

    if len(percent_styles) > 1:
        warnings.add(
            "mixed_percent_format",
            message_en=(
                f"{table.code}: completion is written in {len(percent_styles)} different ways "
                f"({', '.join(sorted(percent_styles))}); all values normalised to per cent."
            ),
            message_az=(
                f"{table.code}: tamamlanma faizi {len(percent_styles)} fərqli formatda yazılıb "
                f"({', '.join(sorted(percent_styles))}); hamısı faizə çevrildi."
            ),
            field_code=table.code,
            sheet=sheet.title,
        )
    return rows


def _table_header_row(sheet: Worksheet, table: TableDef) -> int | None:
    """The ``№`` row that follows the block title — found by text, so rows may move.

    The title must *start* with the marker. Matching anywhere in the cell would catch the
    sheet's own subtitle ("Son 5 ildə tamamlanmış və hazırda davam edən layihələrin
    siyahısı"), which mentions both blocks and sits above the first of them.
    """
    marker = table.marker.casefold()
    found = False
    for row in range(1, sheet.max_row + 1):
        text = clean_text(sheet[f"{CODE_COL}{row}"].value)
        if text is None:
            continue
        if not found:
            found = text.casefold().startswith(marker)
            continue
        if text == "№":
            return row
    return None


# --------------------------------------------------------------------------------------
# Document checklist
# --------------------------------------------------------------------------------------


def _parse_documents(workbook: Any, result: ParsedApplication, warnings: Warnings) -> None:
    sheet = sheet_by_index(workbook, DOCUMENTS_SHEET_INDEX)
    if sheet is None:
        warnings.add(
            "missing_sheet",
            message_en="Document checklist sheet is missing from the workbook.",
            message_az="Sənəd siyahısı vərəqəsi faylda yoxdur.",
            severity="error",
        )
        return

    rows = code_row_index(sheet, CODE_COL)
    notes: list[dict[str, Any]] = []
    for code, definition in DOCUMENT_CATALOG.items():
        row = rows.get(code)
        if row is None:
            result.documents[code] = "missing"
            continue
        word = clean_text(sheet.cell(row=row, column=5).value)
        status = DOCUMENT_STATUS_WORDS.get(word.casefold(), "missing") if word else "missing"
        result.documents[code] = status
        notes.append(
            {
                "code": code,
                "name_az": definition.name_az,
                "name_en": definition.name_en,
                "mandatory": definition.mandatory,
                "status": status,
                "note": clean_text(sheet.cell(row=row, column=6).value),
            }
        )
        if status == "missing":
            warnings.add(
                "document_status_missing",
                message_en=(
                    f"Document {code} ({definition.name_en}) has no status in the checklist."
                ),
                message_az=f"{code} sənədi ({definition.name_az}) üçün status göstərilməyib.",
                severity="error" if definition.mandatory else "warning",
                field_code=code,
                sheet=sheet.title,
                cell=f"E{row}",
            )
    result.document_details = notes

    declared = _declared_document_total(sheet)
    ready = sum(1 for status in result.documents.values() if status == "uploaded")
    if declared is not None and ready < declared:
        warnings.add(
            "document_status_missing",
            message_en=(
                f"The checklist declares {declared} documents but only {ready} are marked ready — "
                f"{declared - ready} statuses are missing."
            ),
            message_az=(
                f"Siyahıda {declared} sənəd elan olunub, "
                f"lakin yalnız {ready} «Hazır» kimi qeyd edilib — "
                f"{declared - ready} status çatışmır."
            ),
            sheet=sheet.title,
            raw_value=f"{ready} / {declared}",
        )


def _declared_document_total(sheet: Worksheet) -> int | None:
    """The checklist's own "N / M" completion cell — M is how many documents it expects."""
    labels = label_row_index(sheet, CODE_COL)
    row = labels.get(normalise_label(DOCUMENT_COMPLETION_LABEL))
    if row is None:
        return None
    for column in range(3, 7):
        text = clean_text(sheet.cell(row=row, column=column).value)
        if text and "/" in text:
            _, _, tail = text.partition("/")
            total = normalise_number(tail)
            if total is not None:
                return int(total)
    return None


# --------------------------------------------------------------------------------------
# Cross-field checks
# --------------------------------------------------------------------------------------


def _fill_vendor(result: ParsedApplication) -> None:
    """Section A is the vendor's own statement and outranks the cover sheet's copy of it."""
    for code, key in VENDOR_FIELD_CODES.items():
        value = result.answers.get(code)
        if value is None:
            continue
        # A year is a year, not 2015.0 — the answers keep the uniform float, the identity
        # record does not, because it is what the vendor row and the UI show.
        if key == "reg_year" and isinstance(value, float) and value.is_integer():
            value = int(value)
        result.vendor[key] = value


def _check_mandatory(result: ParsedApplication, warnings: Warnings) -> None:
    for code in MANDATORY_FIELD_CODES:
        if result.answers.get(code) is None:
            definition = FIELD_CATALOG[code]
            warnings.add(
                "mandatory_cell_empty",
                message_en=f"{code} ({definition.name_en}) is mandatory and was left empty.",
                message_az=f"{code} ({definition.name_az}) məcburidir və boş buraxılıb.",
                severity="error",
                field_code=code,
            )


def _reference_date(meta: dict[str, Any]) -> date | None:
    """The date the form was issued, falling back to its deadline."""
    for key in ("issued_on", "due_on"):
        parsed = parse_iso(meta.get(key))
        if parsed is not None:
            return parsed
    return None


def _check_tax_clearance(result: ParsedApplication, warnings: Warnings) -> None:
    """A-05 is valid for three months from its issue date (brief §1.5, spec §6.1)."""
    issued = parse_iso(result.answers.get("A.16"))
    reference = _reference_date(result.meta)
    if issued is None or reference is None:
        return
    age = months_between(issued, reference)
    if age < TAX_CLEARANCE_VALID_MONTHS:
        return
    warnings.add(
        "stale_certificate",
        message_en=(
            f"A.16: the tax clearance certificate is dated {issued.isoformat()}, "
            f"{age} months before the application date {reference.isoformat()} — "
            f"document A-05 must be no older than {TAX_CLEARANCE_VALID_MONTHS} months."
        ),
        message_az=(
            f"A.16: vergi borcsuzluğu arayışı {issued.isoformat()} tarixlidir — "
            f"müraciət tarixindən ({reference.isoformat()}) {age} ay əvvəl. "
            f"A-05 sənədi {TAX_CLEARANCE_VALID_MONTHS} aydan köhnə ola bilməz."
        ),
        severity="error",
        field_code="A.16",
        raw_value=issued.isoformat(),
    )
