"""Small openpyxl helpers shared by the two parsers.

Both files address rows by what column B says, so every lookup here builds an index of
``column B text -> row number`` once and then reads cells by code. That is what lets a sheet
grow a row without moving an answer (spec §6.1).
"""

from __future__ import annotations

import re
from typing import Any

from openpyxl.utils import column_index_from_string
from openpyxl.worksheet.worksheet import Worksheet

#: A field or document code as it appears in column B: ``A.1``, ``B.13``, ``A-05``, ``C.t1``,
#: a bare group letter on the workbook's totals rows, and the two verdict rows ``KO`` / ``✓``.
CODE_RE = re.compile(r"^(?:KO|[A-HΣ✓](?:\.(?:t?\d{1,2})|-\d{2})?)$")

#: Sheet names start with their position ("4. C. Texniki Təcrübə").
_SHEET_INDEX_RE = re.compile(r"^\s*(\d+)\s*\.")


def sheet_by_index(workbook: Any, index: int) -> Worksheet | None:
    """The sheet whose name starts with ``index``, or ``None``.

    Resolving by the leading number rather than the whole title means a renamed sheet
    ("4. C. Texniki Təcrübə və Layihələr") still parses, while a reordered workbook does
    not silently read the wrong sheet.
    """
    for sheet in workbook.worksheets:
        match = _SHEET_INDEX_RE.match(str(sheet.title))
        if match is not None and int(match.group(1)) == index:
            return sheet
    return None


def normalise_label(text: object) -> str:
    """Row labels differ by trailing colons, case and stray spaces; this ignores all three."""
    if text is None:
        return ""
    return " ".join(str(text).split()).rstrip(":").casefold()


def code_row_index(sheet: Worksheet, column: str) -> dict[str, int]:
    """``{code: row}`` for every code-looking cell in ``column``, first occurrence winning."""
    index: dict[str, int] = {}
    col = column_index_from_string(column)
    for row in range(1, sheet.max_row + 1):
        value = sheet.cell(row=row, column=col).value
        if not isinstance(value, str):
            continue
        code = value.strip()
        if CODE_RE.match(code) and code not in index:
            index[code] = row
    return index


def label_row_index(sheet: Worksheet, column: str) -> dict[str, int]:
    """``{normalised label: row}`` for every text cell in ``column``."""
    index: dict[str, int] = {}
    col = column_index_from_string(column)
    for row in range(1, sheet.max_row + 1):
        label = normalise_label(sheet.cell(row=row, column=col).value)
        if label and label not in index:
            index[label] = row
    return index
