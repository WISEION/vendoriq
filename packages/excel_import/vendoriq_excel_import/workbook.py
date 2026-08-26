"""Reading a Rev4-style scoring workbook — one column per participant.

The workbook is the officers' side of the process: sheet 2 carries each vendor's profile,
sheet 3 the raw indicators they transcribed from the application, sheet 4 the points the
sheet's own formulas produced, and sheet 5 the summary the commission signs. This parser
reads all four so that the engine port can be checked against what the workbook actually
computed — ``sheet_total`` and ``sheet_decision`` are the acceptance fixture for phase 1A.

Rows are addressed by the code in column B; participants are the columns that carry a name
on the profile sheet, so an unused participant column is not a vendor with no data.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from .catalog import (
    CODE_COL,
    WORKBOOK_ANSWER_FIRST_COL,
    WORKBOOK_DECISION_CODE,
    WORKBOOK_GROUP_CODES,
    WORKBOOK_KO_CODE,
    WORKBOOK_POINTS_SHEET_INDEX,
    WORKBOOK_PROFILE_FIRST_COL,
    WORKBOOK_PROFILE_LABELS,
    WORKBOOK_PROFILE_SHEET_INDEX,
    WORKBOOK_RAW_SHEET_INDEX,
    WORKBOOK_TOTAL_CODE,
)
from .normalise import ImportWarning, Warnings, split_multi_value
from .sheets import code_row_index, label_row_index, normalise_label, sheet_by_index

#: Criterion codes of the Rev4 answer sheet, in sheet order.
RAW_INDICATOR_CODES: tuple[str, ...] = (
    "A.1", "A.2", "A.3", "A.4",
    "B.1", "B.2", "B.3", "B.4",
    "C.1", "C.2", "C.3", "C.4",
    "D.1", "D.2", "D.3",
    "E.1", "E.2", "E.3", "E.4",
    "F.1", "F.2", "F.3",
    "G.1", "G.2",
)  # fmt: skip


@dataclass(slots=True)
class WorkbookVendor:
    """One participant column of a scoring workbook.

    Profile values keep the workbook's own types — ``voen`` is an integer for a vendor with
    one VÖEN and the original string for the one that merged two companies — because this
    object has to reproduce ``seed/vendors_seed.json`` byte for byte. The parsed-out list of
    values lives in ``voen_values`` instead, next to the warning that explains it.
    """

    name: str
    voen: Any = None
    reg_year: Any = None
    address: str | None = None
    contact: str | None = None
    position: str | None = None
    phone: str | None = None
    email: str | None = None
    website: str | None = None
    staff: Any = None
    engineers: Any = None
    #: Raw indicator code -> the officer's number or 0–3 rubric value; ``None`` when blank.
    raw: dict[str, float | int | None] = field(default_factory=dict)
    #: Points the workbook's formulas produced, per criterion.
    points: dict[str, float] = field(default_factory=dict)
    #: Points per group (A–G), as the workbook rounded them.
    groups: dict[str, float] = field(default_factory=dict)
    sheet_total: float | None = None
    sheet_ko: str | None = None
    sheet_decision: str | None = None
    #: Both VÖENs when one cell held two ("1400915571 / 7200482051").
    voen_values: list[str] | None = None
    warnings: list[ImportWarning] = field(default_factory=list)

    def to_seed_row(self, index: int) -> dict[str, Any]:
        """The row shape of ``seed/vendors_seed.json`` (``V01`` … ``V13``)."""
        return {
            "id": f"V{index:02d}",
            "name": self.name,
            "voen": self.voen,
            "regYear": self.reg_year,
            "address": self.address,
            "contact": self.contact,
            "position": self.position,
            "phone": self.phone,
            "email": self.email,
            "website": self.website,
            "staff": self.staff,
            "engineers": self.engineers,
            "raw": dict(self.raw),
            "sheetTotal": self.sheet_total,
            "sheetKO": self.sheet_ko,
            "sheetDecision": self.sheet_decision,
        }

    def as_dict(self) -> dict[str, Any]:
        """A JSON-serialisable view — what the CLI prints."""
        return {
            "name": self.name,
            "voen": self.voen,
            "voen_values": self.voen_values,
            "reg_year": self.reg_year,
            "address": self.address,
            "contact": self.contact,
            "position": self.position,
            "phone": self.phone,
            "email": self.email,
            "website": self.website,
            "staff": self.staff,
            "engineers": self.engineers,
            "raw": dict(self.raw),
            "points": dict(self.points),
            "groups": dict(self.groups),
            "sheet_total": self.sheet_total,
            "sheet_ko": self.sheet_ko,
            "sheet_decision": self.sheet_decision,
            "warnings": [w.as_dict() for w in self.warnings],
        }


def parse_scoring_workbook(path: Path | str) -> list[WorkbookVendor]:
    """Read a Rev4-style scoring workbook and return one entry per participant.

    A blank template (the Rev1 fixture) has no named participants and returns ``[]``.
    """
    source = Path(path)
    workbook = load_workbook(source, data_only=True, read_only=False)
    try:
        profile = sheet_by_index(workbook, WORKBOOK_PROFILE_SHEET_INDEX)
        if profile is None:
            return []
        labels = label_row_index(profile, CODE_COL)
        name_row = labels.get(normalise_label("Şirkətin tam adı"))
        if name_row is None:
            return []

        columns = _participant_columns(profile, name_row)
        raw_sheet = sheet_by_index(workbook, WORKBOOK_RAW_SHEET_INDEX)
        points_sheet = sheet_by_index(workbook, WORKBOOK_POINTS_SHEET_INDEX)

        vendors: list[WorkbookVendor] = []
        for offset, column in enumerate(columns):
            name = str(profile.cell(row=name_row, column=column).value)
            vendor = WorkbookVendor(name=name)
            warnings = Warnings()
            _read_profile(profile, labels, column, vendor, warnings)
            answer_column = WORKBOOK_ANSWER_FIRST_COL + offset
            if raw_sheet is not None:
                vendor.raw = _read_raw(raw_sheet, answer_column)
            if points_sheet is not None:
                _read_points(points_sheet, answer_column, vendor)
            vendor.warnings = warnings.items
            vendors.append(vendor)
        return vendors
    finally:
        workbook.close()


def to_seed_rows(vendors: list[WorkbookVendor]) -> list[dict[str, Any]]:
    """``seed/vendors_seed.json`` as the seed CLI and the fixture test expect it."""
    return [vendor.to_seed_row(index) for index, vendor in enumerate(vendors, start=1)]


# --------------------------------------------------------------------------------------


def _participant_columns(profile: Worksheet, name_row: int) -> list[int]:
    """Columns that carry a participant name. The template leaves them blank or at ``0``."""
    columns: list[int] = []
    for column in range(WORKBOOK_PROFILE_FIRST_COL, profile.max_column + 1):
        value = profile.cell(row=name_row, column=column).value
        if isinstance(value, str) and value.strip():
            columns.append(column)
    return columns


def _read_profile(
    profile: Worksheet,
    labels: dict[str, int],
    column: int,
    vendor: WorkbookVendor,
    warnings: Warnings,
) -> None:
    """Profile cells are copied verbatim — including the line breaks inside an address.

    The workbook is the record of what the officers typed, and ``seed/vendors_seed.json``
    is that record; tidying the whitespace here would silently rewrite it.
    """
    for label, key in WORKBOOK_PROFILE_LABELS.items():
        if key == "name":
            continue
        row = labels.get(normalise_label(label))
        if row is None:
            continue
        cell = profile.cell(row=row, column=column)
        raw = cell.value
        if raw is None or (isinstance(raw, str) and not raw.strip()):
            continue
        setattr(vendor, key, raw)
        if key == "voen":
            parts = split_multi_value(raw)
            if parts is not None:
                vendor.voen_values = parts
                warnings.add(
                    "multi_value_cell",
                    message_en=(
                        f"{vendor.name}: the VÖEN cell holds {len(parts)} numbers "
                        f"({', '.join(parts)}) — the vendor was registered twice."
                    ),
                    message_az=(
                        f"{vendor.name}: VÖEN xanasında {len(parts)} nömrə var "
                        f"({', '.join(parts)}) — şirkət iki dəfə qeydiyyatdan keçib."
                    ),
                    field_code="voen",
                    sheet=profile.title,
                    cell=cell.coordinate,
                    raw_value=raw,
                )


def _read_raw(sheet: Worksheet, column: int) -> dict[str, float | int | None]:
    rows = code_row_index(sheet, CODE_COL)
    values: dict[str, float | int | None] = {}
    for code in RAW_INDICATOR_CODES:
        row = rows.get(code)
        if row is None:
            values[code] = None
            continue
        value = sheet.cell(row=row, column=column).value
        values[code] = (
            value if isinstance(value, int | float) and not isinstance(value, bool) else None
        )
    return values


def _read_points(sheet: Worksheet, column: int, vendor: WorkbookVendor) -> None:
    rows = code_row_index(sheet, CODE_COL)
    for code in RAW_INDICATOR_CODES:
        row = rows.get(code)
        if row is None:
            continue
        value = sheet.cell(row=row, column=column).value
        if isinstance(value, int | float) and not isinstance(value, bool):
            vendor.points[code] = float(value)
    for group in WORKBOOK_GROUP_CODES:
        row = rows.get(group)
        if row is None:
            continue
        value = sheet.cell(row=row, column=column).value
        if isinstance(value, int | float) and not isinstance(value, bool):
            vendor.groups[group] = float(value)

    total_row = rows.get(WORKBOOK_TOTAL_CODE)
    if total_row is not None:
        total = sheet.cell(row=total_row, column=column).value
        if isinstance(total, int | float) and not isinstance(total, bool):
            vendor.sheet_total = total
    vendor.sheet_ko = _verdict(sheet, rows.get(WORKBOOK_KO_CODE), column)
    vendor.sheet_decision = _verdict(sheet, rows.get(WORKBOOK_DECISION_CODE), column)


def _verdict(sheet: Worksheet, row: int | None, column: int) -> str | None:
    """The KO / decision cell, verbatim: "Keçdi ✓", "KO — RƏDD"."""
    if row is None:
        return None
    value = sheet.cell(row=row, column=column).value
    return value if isinstance(value, str) and value.strip() else None
